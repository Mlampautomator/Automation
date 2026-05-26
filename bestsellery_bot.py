#!/usr/bin/env python3
"""
Bestsellery Bot — zbiera bestsellery z 4 sklepów oświetleniowych.
Pobiera: Symbol/Indeks produktu, Cena katalogowa (przed promocją), Cena promocyjna

Sklepy: polskielampy.pl, kinkiecik.pl, bajkowelampy.pl, tomix.pl

Wymagania: pip install aiohttp beautifulsoup4 openpyxl lxml
Uruchomienie: python bestsellery_bot.py
"""

import re, time, asyncio, aiohttp, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ── Konfiguracja ────────────────────────────────────────────────────────────────
PLIK_WYNIKI = "bestsellery_wyniki.xlsx"
MAX_STRONY  = 5      # max stron paginacji per sklep
CONCURRENCY = 5
TIMEOUT     = 30
OPOZNIENIE  = 0.4
MAX_RETRY   = 3

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pl-PL,pl;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Konfiguracja sklepów ────────────────────────────────────────────────────────
SHOPS = [
    {
        "name":     "PolskieLampy",
        "url":      "https://polskielampy.pl/bestsellery.html",
        "base":     "https://polskielampy.pl",
        "platform": "shoper_a",
    },
    {
        "name":     "Kinkiecik",
        "url":      "https://kinkiecik.pl/hity.html",
        "base":     "https://kinkiecik.pl",
        "platform": "shoper_a",
    },
    {
        "name":         "BajkoweLampy",
        "url":          "https://bajkowelampy.pl/Bestseller-sbestseller-zul.html",
        "base":         "https://bajkowelampy.pl",
        "platform":     "bajkowelampy",
        "link_pattern": r"product-zul-\d+",
    },
    {
        "name":         "Tomix",
        "url":          "https://www.tomix.pl/bestsellery",
        "base":         "https://www.tomix.pl",
        "platform":     "tomix",
        "link_pattern": r"/product-pol-\d+",
    },
]

# ── HTTP ─────────────────────────────────────────────────────────────────────────
async def fetch(session: aiohttp.ClientSession, url: str):
    for attempt in range(1, MAX_RETRY + 1):
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=TIMEOUT)) as r:
                if r.status == 429:
                    wait = 2 ** attempt
                    print(f"  429 rate-limit, czekam {wait}s...")
                    await asyncio.sleep(wait)
                    continue
                r.raise_for_status()
                return BeautifulSoup(await r.text(), "lxml")
        except asyncio.TimeoutError:
            print(f"  timeout (próba {attempt}/{MAX_RETRY}): {url}")
            await asyncio.sleep(attempt)
        except Exception as e:
            print(f"  błąd (próba {attempt}/{MAX_RETRY}) {url}: {e}")
            await asyncio.sleep(attempt)
    return None

# ── Pomocnicze ───────────────────────────────────────────────────────────────────
def wyciagnij_cene(text: str) -> str:
    """Wyciąga cenę z tekstu: '89,00 zł' → '89,00', '17 083,00 zł' → '17083,00'"""
    m = re.search(r"(\d[\d\s]*[,.]\d{2})", str(text))
    return m.group(1).replace("\xa0", "").replace(" ", "") if m else ""

def normalizuj_cene(val: str) -> str:
    """Zamienia '89.00' → '89,00'. Zostawia '89,00' bez zmian."""
    val = (val or "").strip()
    if re.match(r"^\d+\.\d{2}$", val):
        return val.replace(".", ",")
    return wyciagnij_cene(val)

# ── Parsery stron produktów ──────────────────────────────────────────────────────

def parsuj_shoper_a(soup: BeautifulSoup):
    """
    polskielampy.pl, kinkiecik.pl — platforma IAI/Shoper
    SKU:          <strong itemprop="mpn">AZ6855</strong>
    Cena promo:   <span itemprop="price" content="119.00">119,00 zł</span>
    Cena katalog: <p id="CenaPoprzednia"><strong content="89.00">89,00 zł</strong></p>
    """
    # Symbol / indeks
    el = soup.find(itemprop="mpn")
    symbol = el.get_text(strip=True) if el else ""

    # Cena promocyjna (aktualna)
    cena_promo = ""
    el = soup.find(itemprop="price")
    if el:
        cena_promo = wyciagnij_cene(el.get_text()) or normalizuj_cene(el.get("content", ""))
    if not cena_promo:
        el = soup.find(id="CenaGlownaProduktuBrutto")
        if el:
            cena_promo = wyciagnij_cene(el.get_text())

    # Cena katalogowa / przed promocją
    cena_kat = ""
    el = soup.select_one("#CenaPoprzednia strong")
    if el:
        cena_kat = normalizuj_cene(el.get("content", "")) or wyciagnij_cene(el.get_text())
    if not cena_kat:
        el = soup.find(id="CenaPoprzednia")
        if el:
            cena_kat = wyciagnij_cene(el.get_text())

    return symbol, cena_promo, cena_kat


def parsuj_bajkowelampy(soup: BeautifulSoup):
    """
    bajkowelampy.pl
    SKU:          <strong itemprop="productID" content="mpn:honice s150">honice s150</strong>
    Cena promo:   <strong id="projector_price_value">...<span class="price">985.00</span>...</strong>
    Cena katalog: <span class="projector_price_srp" id="projector_price_srp">985,00 zł</span>
    """
    # Symbol / indeks
    el = soup.find(itemprop="productID")
    symbol = el.get_text(strip=True) if el else ""

    # Cena promocyjna — span.price wewnątrz #projector_price_value (format "985.00")
    cena_promo = ""
    el = soup.select_one("#projector_price_value span.price")
    if el:
        cena_promo = normalizuj_cene(el.get_text(strip=True))
    if not cena_promo:
        el = soup.find(id="projector_price_value")
        if el:
            cena_promo = normalizuj_cene(wyciagnij_cene(el.get_text()))

    # Cena katalogowa
    el = soup.find(id="projector_price_srp")
    cena_kat = wyciagnij_cene(el.get_text()) if el else ""

    return symbol, cena_promo, cena_kat


def parsuj_tomix(soup: BeautifulSoup):
    """
    tomix.pl
    SKU:          <div class="dictionary__param">
                    <span class="dictionary__name_txt">Kod produktu</span>
                    <span class="dictionary__value_txt">MLP5739</span>
                  </div>
    Cena promo:   <strong class="projector_prices__price" id="projector_price_value">96,60 zł</strong>
    Cena katalog: <strong class="projector_prices__srp" id="projector_price_srp">115,00 zł</strong>
    """
    # Symbol / indeks — szukaj wiersza "Kod produktu" w tabeli parametrów
    symbol = ""
    for row in soup.select(".dictionary__param"):
        name_el = row.select_one(".dictionary__name_txt")
        val_el  = row.select_one(".dictionary__value_txt")
        if name_el and val_el and "kod produktu" in name_el.get_text(strip=True).lower():
            symbol = val_el.get_text(strip=True)
            break

    # Cena promocyjna
    el = soup.find(id="projector_price_value")
    cena_promo = wyciagnij_cene(el.get_text()) if el else ""

    # Cena katalogowa
    el = soup.find(id="projector_price_srp")
    cena_kat = wyciagnij_cene(el.get_text()) if el else ""

    return symbol, cena_promo, cena_kat

# ── Pobieranie listy produktów ───────────────────────────────────────────────────
async def get_product_links(session: aiohttp.ClientSession, shop: dict) -> list:
    """Zbiera URLe produktów ze stron z bestsellerami (z paginacją)."""
    base    = shop["base"]
    url     = shop["url"]
    seen    = set()
    links   = []

    for page_nr in range(1, MAX_STRONY + 1):
        print(f"    strona {page_nr}: {url}")
        soup = await fetch(session, url)
        if not soup:
            break

        if shop["platform"] == "shoper_a":
            # IAI/Shoper — każdy produkt w div.ProdCena z linkiem w h3 a
            for card in soup.select("div.ProdCena"):
                a = card.select_one("h3 a")
                if not a:
                    continue
                href = a.get("href", "")
                if not href:
                    continue
                if not href.startswith("http"):
                    href = base + href
                if href not in seen:
                    seen.add(href)
                    links.append(href)
        else:
            # bajkowelampy / tomix — filtruj linki po wzorcu URL produktu
            pattern = shop.get("link_pattern", r"product-")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if not re.search(pattern, href):
                    continue
                if not href.startswith("http"):
                    href = base + "/" + href.lstrip("/")
                if href not in seen:
                    seen.add(href)
                    links.append(href)

        # Paginacja (standard rel=next)
        next_el = soup.find("link", rel="next") or soup.find("a", rel="next")
        if not next_el:
            break
        next_href = next_el.get("href", "")
        if not next_href:
            break
        if not next_href.startswith("http"):
            next_href = base + "/" + next_href.lstrip("/")
        url = next_href
        await asyncio.sleep(OPOZNIENIE)

    return links

# ── Pobieranie szczegółów produktu ───────────────────────────────────────────────
async def scrape_product(session: aiohttp.ClientSession, url: str, shop: dict, semafor: asyncio.Semaphore):
    async with semafor:
        soup = await fetch(session, url)
        await asyncio.sleep(OPOZNIENIE)

    if not soup:
        return None

    h1 = soup.find("h1")
    nazwa = h1.get_text(strip=True) if h1 else ""

    platform = shop["platform"]
    if platform == "shoper_a":
        symbol, cena_promo, cena_kat = parsuj_shoper_a(soup)
    elif platform == "bajkowelampy":
        symbol, cena_promo, cena_kat = parsuj_bajkowelampy(soup)
    else:
        symbol, cena_promo, cena_kat = parsuj_tomix(soup)

    return {
        "Sklep":           shop["name"],
        "Nazwa":           nazwa,
        "Symbol":          symbol,
        "Cena_katalogowa": cena_kat,
        "Cena_promocyjna": cena_promo,
        "URL":             url,
        "Data":            datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

# ── Zapis do Excel ───────────────────────────────────────────────────────────────
def save_to_excel(data: list, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bestsellery"

    KEYS     = ["Sklep", "Nazwa", "Symbol", "Cena_katalogowa", "Cena_promocyjna", "URL", "Data"]
    HEADERS  = ["Sklep", "Nazwa produktu", "Symbol / Indeks", "Cena przed promocją (zł)",
                "Cena promocyjna (zł)", "Link do produktu", "Data pobrania"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for row_data in data:
        ws.append([row_data.get(k, "") for k in KEYS])

    col_widths = [15, 55, 20, 26, 24, 65, 18]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Zamroź nagłówek
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\nZapisano {len(data)} produktow -> {filename}")

# ── Main ─────────────────────────────────────────────────────────────────────────
async def main_async():
    print("=" * 55)
    print("  Bestsellery Bot")
    print("=" * 55)
    semafor  = asyncio.Semaphore(CONCURRENCY)
    all_data = []
    t_start  = time.time()

    connector = aiohttp.TCPConnector(limit=CONCURRENCY, ttl_dns_cache=300)
    async with aiohttp.ClientSession(headers=HEADERS, connector=connector) as session:

        for shop in SHOPS:
            print(f"\n[{shop['name']}]")

            t0    = time.time()
            links = await get_product_links(session, shop)
            print(f"  Znaleziono {len(links)} produktów ({time.time()-t0:.1f}s)")

            if not links:
                print("  Brak produktów — pomijam.")
                continue

            tasks  = [asyncio.create_task(scrape_product(session, url, shop, semafor)) for url in links]
            wyniki = []
            t1     = time.time()

            for i, coro in enumerate(asyncio.as_completed(tasks), 1):
                result = await coro
                if result:
                    wyniki.append(result)
                if i % 10 == 0 or i == len(tasks):
                    elapsed = time.time() - t1
                    rps = i / elapsed if elapsed > 0 else 0
                    print(f"  [{i}/{len(tasks)}] {elapsed:.1f}s | {rps:.1f} prod/s")

            print(f"  Zebrano: {len(wyniki)}/{len(links)} produktów")
            all_data.extend(wyniki)

    elapsed = time.time() - t_start
    print(f"\n{'='*55}")
    print(f"  Łącznie: {len(all_data)} produktów w {elapsed:.1f}s")
    print(f"{'='*55}")

    if all_data:
        save_to_excel(all_data, PLIK_WYNIKI)
        try:
            import os
            os.startfile(PLIK_WYNIKI)
            print("Otwieranie pliku wynikowego...")
        except Exception:
            pass
    else:
        print("Brak danych do zapisania.")

def main():
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(main_async())

if __name__ == "__main__":
    main()
