"""
Scraper wyprzedażowy lamp - polskielampy.pl i konkurenci
Odczytuje linki z pliku links.xlsx, scrapuje produkty i zapisuje wyniki.

Wymagania:
    pip install requests beautifulsoup4 openpyxl
"""

import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ────────────────────────────────────────────────
# KONFIGURACJA
# ────────────────────────────────────────────────
PLIK_LINKI   = "links.xlsx"
PLIK_WYNIKI  = "wyniki_wyprzedaz.xlsx"
OPOZNIENIE   = 1.5       # sekundy między requestami (nie przeciążaj serwera)
MAX_STRONY   = 20        # maks. liczba podstron paginacji na jeden sklep

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pl-PL,pl;q=0.9",
}


# ────────────────────────────────────────────────
# ODCZYT LINKÓW Z EXCELA
# ────────────────────────────────────────────────
def wczytaj_linki(plik: str) -> list[dict]:
    df = pd.read_excel(plik, dtype=str)
    df.columns = df.columns.str.strip()
    aktywne = df[df["Aktywny"].str.upper() == "TAK"]
    return aktywne[["Sklep", "URL_wyprzedazy"]].to_dict("records")


# ────────────────────────────────────────────────
# POMOCNICZE
# ────────────────────────────────────────────────
def pobierz_strone(url: str) -> BeautifulSoup | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        return BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"  ✗ Błąd pobierania {url}: {e}")
        return None


def oczysc_cene(tekst: str) -> str:
    """Zwraca czystą cenę bez nadmiarowych białych znaków."""
    return re.sub(r"\s+", " ", tekst).strip()


# ────────────────────────────────────────────────
# PARSER: polskielampy.pl
# ────────────────────────────────────────────────
def scrape_polskielampy(url_start: str, sklep: str) -> list[dict]:
    """
    Struktura produktu na polskielampy.pl:
      <div class="produkt-box"> (lub "product-box")
        <a class="name">  → nazwa + link
        .cena-regular     → cena regularna (przekreślona)
        .cena-promocyjna  → cena promocyjna / wyprzedażowa
        span.discount     → procent rabatu (np. "-30%")
    """
    produkty = []

    for nr_strony in range(1, MAX_STRONY + 1):
        url = url_start if nr_strony == 1 else f"{url_start}?strona={nr_strony}"
        print(f"  → Pobieranie strony {nr_strony}: {url}")

        soup = pobierz_strone(url)
        if soup is None:
            break

        # Szukamy kontenerów produktów
        boxy = soup.select("div.produkt-box, div.product-box, li.produkt")
        if not boxy:
            # Próba alternatywnego selektora
            boxy = soup.select("[class*='produkt']")

        if not boxy:
            print(f"  ℹ Brak produktów na stronie {nr_strony} – koniec paginacji")
            break

        for box in boxy:
            try:
                # Nazwa i URL produktu
                link_tag = box.select_one("a.name, a[class*='name'], h2 a, h3 a")
                if not link_tag:
                    continue
                nazwa = link_tag.get_text(strip=True)
                url_produktu = link_tag.get("href", "")
                if url_produktu and not url_produktu.startswith("http"):
                    url_produktu = "https://polskielampy.pl" + url_produktu

                # Ceny
                el_stara = box.select_one(".cena-regular, .price-old, [class*='regular'], del")
                el_nowa  = box.select_one(".cena-promocyjna, .price-promo, [class*='promo'], .cena-aktualna")
                el_rabat = box.select_one(".discount, .rabat, [class*='discount'], [class*='rabat']")

                cena_stara = oczysc_cene(el_stara.get_text()) if el_stara else ""
                cena_nowa  = oczysc_cene(el_nowa.get_text())  if el_nowa  else ""
                rabat      = oczysc_cene(el_rabat.get_text()) if el_rabat else ""

                # Jeśli brak ceny promocyjnej, szukaj po prostu ceny
                if not cena_nowa:
                    el_cena = box.select_one(".cena, .price, [class*='cena']")
                    if el_cena:
                        cena_nowa = oczysc_cene(el_cena.get_text())

                if not nazwa or not cena_nowa:
                    continue

                produkty.append({
                    "Sklep":        sklep,
                    "Nazwa":        nazwa,
                    "Cena_stara":   cena_stara,
                    "Cena_nowa":    cena_nowa,
                    "Rabat":        rabat,
                    "URL_produktu": url_produktu,
                    "Data_scrape":  datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
            except Exception as e:
                print(f"    Błąd parsowania produktu: {e}")
                continue

        # Sprawdzenie czy jest następna strona
        nastepna = soup.select_one("a.next, a[rel='next'], .pagination .next a, a[class*='nastepna']")
        if not nastepna:
            print(f"  ✓ Koniec stron ({nr_strony} stron)")
            break

        time.sleep(OPOZNIENIE)

    return produkty


# ────────────────────────────────────────────────
# PARSER UNIWERSALNY (dla nieznanych stron)
# ────────────────────────────────────────────────
def scrape_uniwersalny(url: str, sklep: str) -> list[dict]:
    """
    Próbuje typowych wzorców e-commerce.
    Dla każdego nowego konkurenta warto napisać dedykowaną funkcję
    (tak jak scrape_polskielampy powyżej) po obejrzeniu jego HTML.
    """
    soup = pobierz_strone(url)
    if soup is None:
        return []

    produkty = []
    # Typowe selektory dla sklepów PrestaShop / WooCommerce / Shoper
    selektory_produktow = [
        ".product-miniature", ".product_list li",
        ".products article", ".woocommerce-loop-product",
        "[class*='product-item']", "[class*='produkt']",
    ]
    boxy = []
    for sel in selektory_produktow:
        boxy = soup.select(sel)
        if boxy:
            break

    for box in boxy:
        try:
            el_nazwa = box.select_one("h2, h3, .product-name, .product-title, [class*='name']")
            el_cena_nowa = box.select_one(".price, .woocommerce-Price-amount, [class*='price']")
            el_cena_stara = box.select_one("del, .regular-price, .price-old, [class*='old']")
            el_link = box.select_one("a[href]")

            nazwa     = el_nazwa.get_text(strip=True)    if el_nazwa     else "?"
            cena_nowa = oczysc_cene(el_cena_nowa.get_text()) if el_cena_nowa else ""
            cena_stara = oczysc_cene(el_cena_stara.get_text()) if el_cena_stara else ""
            link_url  = el_link.get("href", "")          if el_link      else ""

            if not cena_nowa:
                continue

            produkty.append({
                "Sklep":        sklep,
                "Nazwa":        nazwa,
                "Cena_stara":   cena_stara,
                "Cena_nowa":    cena_nowa,
                "Rabat":        "",
                "URL_produktu": link_url,
                "Data_scrape":  datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
        except Exception:
            continue

    return produkty


# ────────────────────────────────────────────────
# ZAPIS DO EXCELA Z FORMATOWANIEM
# ────────────────────────────────────────────────
def zapisz_wyniki(produkty: list[dict], plik: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Wyniki"

    naglowki = ["Sklep", "Nazwa", "Cena stara", "Cena nowa", "Rabat", "URL produktu", "Data scrape"]
    ws.append(naglowki)

    # Styl nagłówka
    nagl_fill = PatternFill("solid", start_color="1F3864")
    thin = Side(style="thin")
    for col, h in enumerate(naglowki, 1):
        cell = ws.cell(1, col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = nagl_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Kolory naprzemienne wierszy
    fill_jasny  = PatternFill("solid", start_color="DCE6F1")
    fill_bialy  = PatternFill("solid", start_color="FFFFFF")

    for i, p in enumerate(produkty):
        row = [
            p["Sklep"], p["Nazwa"], p["Cena_stara"],
            p["Cena_nowa"], p["Rabat"], p["URL_produktu"], p["Data_scrape"]
        ]
        ws.append(row)
        r = ws.max_row
        fill = fill_jasny if i % 2 == 0 else fill_bialy
        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if col == 6:  # URL
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    # Szerokości kolumn
    szerokosci = [18, 50, 14, 14, 10, 55, 18]
    for col, w in enumerate(szerokosci, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Arkusz podsumowanie
    ws2 = wb.create_sheet("Podsumowanie")
    ws2.append(["Sklep", "Liczba produktów"])
    ws2.cell(1,1).font = Font(bold=True, name="Arial")
    ws2.cell(1,2).font = Font(bold=True, name="Arial")
    sklepy = {}
    for p in produkty:
        sklepy[p["Sklep"]] = sklepy.get(p["Sklep"], 0) + 1
    for s, cnt in sklepy.items():
        ws2.append([s, cnt])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    wb.save(plik)
    print(f"\n✓ Zapisano {len(produkty)} produktów → {plik}")


# ────────────────────────────────────────────────
# GŁÓWNA PĘTLA
# ────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  SCRAPER WYPRZEDAŻY LAMP")
    print("=" * 55)

    try:
        linki = wczytaj_linki(PLIK_LINKI)
    except FileNotFoundError:
        print(f"✗ Nie znaleziono pliku {PLIK_LINKI}")
        return

    print(f"Wczytano {len(linki)} aktywnych linków.\n")
    wszystkie = []

    for wpis in linki:
        sklep = wpis["Sklep"]
        url   = wpis["URL_wyprzedazy"]
        print(f"[{sklep}] → {url}")

        # Dobierz parser do sklepu
        if "polskielampy.pl" in url:
            produkty = scrape_polskielampy(url, sklep)
        else:
            # Dla nowych sklepów użyj parsera uniwersalnego
            # lub dopisz nową funkcję scrape_NAZWA(url, sklep)
            produkty = scrape_uniwersalny(url, sklep)

        print(f"  Znaleziono: {len(produkty)} produktów")
        wszystkie.extend(produkty)
        time.sleep(OPOZNIENIE)

    if wszystkie:
        zapisz_wyniki(wszystkie, PLIK_WYNIKI)
    else:
        print("\nℹ Brak produktów do zapisania.")


if __name__ == "__main__":
    main()
