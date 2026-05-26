#!/usr/bin/env python3
"""
Aplikacja do pobierania obrazów produktów z mlamp.pl
Czyta listę URL-i z Excel, pobiera obrazy i zapisuje wyniki.
"""

import subprocess
import sys
import os
import time
import re
from urllib.parse import urljoin
from pathlib import Path

# Auto-zainstaluj wymagane pakiety
def install_packages():
    packages = ['openpyxl', 'requests', 'beautifulsoup4']
    for package in packages:
        try:
            __import__(package.replace('-', ''))
        except ImportError:
            print(f"Instalowanie {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])
    print("✓ Wszystkie pakiety są gotowe!\n")

try:
    install_packages()
except Exception as e:
    print(f"❌ Błąd instalacji pakietów: {e}")
    sys.exit(1)

# Importy
from openpyxl import Workbook, load_workbook
import logging
import requests
from bs4 import BeautifulSoup

# Konfiguracja logowania
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)


class ProductImageScraper:
    def __init__(self):
        pass

    def extract_product_id(self, url: str) -> str:
        """Wyodrębnia ID produktu z URL-u"""
        match = re.search(r'-(\d{5,6})(?:-|$)', url)
        if match:
            return match.group(1)
        return None

    def get_product_name_requests(self, url: str) -> str:
        """Pobiera nazwę produktu przez requests"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }

            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')

            # Szukaj tytułu w różnych miejscach
            selectors = ['h1', '.product-title h1', '[itemprop="name"]', '.productTitle']
            for selector in selectors:
                element = soup.select_one(selector)
                if element:
                    text = element.get_text().strip()
                    if text:
                        return text

            return "Nieznana nazwa"
        except Exception as e:
            logger.debug(f"Info: Błąd pobierania nazwy: {e}")
            return "Nieznana nazwa"

    def get_product_images_requests(self, url: str) -> list:
        """Pobiera linki do obrazów produktu przez requests"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }

            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')
            images = []

            # Znajdź wszystkie obrazki
            for img in soup.find_all('img'):
                src = img.get('src') or img.get('data-src') or img.get('data-original')
                if src:
                    # Bardziej agresywne szukanie - wszystkie obrazy z http
                    if 'http' in src and not any(x in src.lower() for x in ['logo', 'icon', 'banner', 'social']):
                        full_url = urljoin(url, src)
                        if full_url not in images:
                            images.append(full_url)

            logger.info(f"✓ Znaleziono {len(images)} obrazów przez requests")
            return images

        except Exception as e:
            logger.error(f"❌ Błąd pobierania obrazów: {e}")
            return []

    def download_images(self, image_urls: list, product_id: str, download_dir: str = "images"):
        """Pobiera obrazy na dysk z nazwami ID-1.jpg, ID-2.jpg, itd."""
        if not image_urls:
            return []

        # Utwórz folder jeśli nie istnieje
        download_path = Path(download_dir)
        download_path.mkdir(exist_ok=True)

        downloaded_files = []
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
            'Accept-Language': 'pl-PL,pl;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }

        for idx, url in enumerate(image_urls, start=1):
            try:
                filename = f"{product_id}-{idx}.jpg"
                filepath = download_path / filename

                # Pobierz obraz z lepszymi nagłówkami
                response = requests.get(url, headers=headers, timeout=30, stream=True, allow_redirects=True)
                response.raise_for_status()
                
                # Sprawdź rozmiar
                content_length = response.headers.get('content-length')
                if content_length and int(content_length) == 0:
                    logger.warning(f"⚠️  Plik pusty: {filename}")
                    continue

                # Zapisz na dysk
                with open(filepath, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                
                # Sprawdź czy plik ma zawartość
                file_size = filepath.stat().st_size
                if file_size == 0:
                    logger.warning(f"⚠️  Plik pusty po zapisaniu: {filename}")
                    filepath.unlink()  # Usuń pusty plik
                    continue

                downloaded_files.append(str(filepath))
                logger.info(f"✓ Pobrano: {filename} ({file_size} bytes)")

            except Exception as e:
                logger.warning(f"❌ Błąd pobierania {url}: {e}")

        return downloaded_files

    def get_product_info(self, url: str) -> dict:
        """Pobiera informacje o produkcie: nazwę i obrazy"""
        try:
            logger.info(f"📍 Pobieranie: {url}")

            # Wyodrębniaj ID produktu
            product_id = self.extract_product_id(url)
            if not product_id:
                logger.warning(f"⚠️  Nie znaleziono ID produktu w URL-u")
                return None

            # Pobierz informacje przez requests
            product_name = self.get_product_name_requests(url)
            image_urls = self.get_product_images_requests(url)

            # Pobierz obrazy na dysk
            downloaded_files = []
            if image_urls:
                downloaded_files = self.download_images(image_urls, product_id)
                image_names = [f"{product_id}-{i+1}" for i in range(len(image_urls))]
                logger.info(f"✓ Pobrano {len(downloaded_files)}/{len(image_urls)} obrazów")
            else:
                logger.warning(f"⚠️  Brak obrazów dla produktu {product_id}")

            return {
                'product_name': product_name,
                'product_id': product_id,
                'image_urls': image_urls,
                'downloaded_files': downloaded_files,
                'image_names': image_names,
                'image_count': len(image_urls)
            }

        except Exception as e:
            logger.error(f"❌ Błąd przetwarzania URL-u: {e}")
            return None


def read_urls_from_excel(filepath: str) -> list:
    """Czyta listę URL-i z pliku Excel"""
    try:
        workbook = load_workbook(filepath)
        worksheet = workbook.active

        urls = []
        for row in worksheet.iter_rows(values_only=True):
            url = row[0] if row[0] else None
            if url and isinstance(url, str):
                url = url.strip()
                if url.startswith('http'):
                    urls.append(url)

        logger.info(f"✓ Wczytano {len(urls)} URL-i z pliku Excel")
        return urls

    except FileNotFoundError:
        logger.error(f"❌ Plik nie znaleziony: {filepath}")
        return []
    except Exception as e:
        logger.error(f"❌ Błąd czytania pliku Excel: {e}")
        return []


def save_results_to_excel(results: list, output_filepath: str):
    """Zapisuje wyniki do pliku Excel"""
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Produkty i Obrazy"

        # Nagłówki
        worksheet['A1'] = "Nazwa produktu"
        worksheet['B1'] = "ID produktu"
        worksheet['C1'] = "Liczba obrazów"
        worksheet['D1'] = "Nazwy obrazów"
        worksheet['E1'] = "Ścieżki plików"
        worksheet['F1'] = "Linki do obrazów"

        # Style nagłówka
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            worksheet[cell].font = header_font
            worksheet[cell].fill = header_fill

        # Dane
        for idx, result in enumerate(results, start=2):
            if result:
                worksheet[f'A{idx}'] = result['product_name']
                worksheet[f'B{idx}'] = result['product_id']
                worksheet[f'C{idx}'] = result['image_count']
                worksheet[f'D{idx}'] = ', '.join(result['image_names']) if result['image_names'] else 'Brak'
                worksheet[f'E{idx}'] = '\n'.join(result['downloaded_files']) if result['downloaded_files'] else 'Brak'
                worksheet[f'F{idx}'] = '\n'.join(result['image_urls']) if result['image_urls'] else 'Brak'

        # Dostosuj szerokość kolumn
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 50
        worksheet.column_dimensions['F'].width = 50

        workbook.save(output_filepath)
        logger.info(f"✓ Wyniki zapisane do: {output_filepath}")

    except Exception as e:
        logger.error(f"❌ Błąd zapisania wyniku: {e}")


def main():
    print("\n" + "="*60)
    print("  🖼️  POBIERACZ OBRAZÓW PRODUKTÓW - MLAMP.PL")
    print("="*60 + "\n")

    # Ścieżki
    script_dir = Path(__file__).parent
    input_file = script_dir / "urls.xlsx"
    output_file = script_dir / "output.xlsx"

    # Sprawdź plik wejściowy
    if not input_file.exists():
        logger.error(f"❌ Plik {input_file} nie istnieje!")
        logger.info("   Utwórz plik Excel z URL-ami produktów w kolumnie A")
        sys.exit(1)

    # Czytaj URL-i
    urls = read_urls_from_excel(str(input_file))
    if not urls:
        logger.error("❌ Brak URL-i w pliku!")
        sys.exit(1)

    logger.info(f"📌 Do przetworzenia: {len(urls)} produktów\n")

    # Inicjalizuj scraper
    scraper = ProductImageScraper()
    results = []

    try:
        for idx, url in enumerate(urls, start=1):
            logger.info(f"\n[{idx}/{len(urls)}]")
            result = scraper.get_product_info(url)
            if result:
                results.append(result)
            time.sleep(1)  # Czekaj między żądaniami

    except KeyboardInterrupt:
        logger.info("\n⏹️  Przerwano przez użytkownika")
    except Exception as e:
        logger.error(f"\n❌ Błąd: {e}")

    # Zapisz wyniki
    logger.info(f"\n✓ Przetworzono: {len(results)}/{len(urls)} produktów")
    save_results_to_excel(results, str(output_file))

    logger.info("\n" + "="*60)
    logger.info("✓ GOTOWE!")
    logger.info("="*60)
    
    # Otwórz plik Excel
    logger.info("\n📂 Otwieranie pliku Excel...\n")
    try:
        import subprocess
        subprocess.Popen(['cmd', '/c', f'start "" "{output_file}"'])
        time.sleep(1)
    except Exception as e:
        logger.warning(f"⚠️  Nie udało się otworzyć Excela: {e}")
        logger.info(f"📄 Plik znajduje się w: {output_file}")


if __name__ == "__main__":
    main()