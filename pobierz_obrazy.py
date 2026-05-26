#!/usr/bin/env python3
"""
Aplikacja do pobierania informacji o obrazach produktów z mlamp.pl
Czyta listę URL-i z Excel, wyodrębnia ID produktu i linki do obrazów,
zapisuje wyniki do nowego pliku Excel.
"""

import subprocess
import sys
import os
import time
import re
from urllib.parse import urlparse, urljoin
from pathlib import Path

# Auto-zainstaluj wymagane pakiety
def install_packages():
    packages = ['openpyxl', 'requests', 'beautifulsoup4']
    for package in packages:
        try:
            __import__(package.replace('-', ''))
        except ImportError:
            print(f"Instalowanie {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])
    print("✓ Wszystkie pakiety są gotowe!\n")

try:
    install_packages()
except Exception as e:
    print(f"❌ Błąd instalacji pakietów: {e}")
    sys.exit(1)

# Importy
from openpyxl import Workbook, load_workbook
import logging
import requests
from pathlib import Path
from bs4 import BeautifulSoup

# Konfiguracja logowania
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)


class ProductImageScraper:
    def __init__(self):
        pass

    def extract_product_id(self, url: str) -> str:
        """Wyodrębnia ID produktu z URL-u"""
        match = re.search(r'-(\d{5,6})(?:-|$)', url)
        if match:
            return match.group(1)
        return None

    def fetch_page_soup(self, url: str):
        """Pobiera stronę produktu i zwraca obiekt BeautifulSoup"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'pl-PL,pl;q=0.9',
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return BeautifulSoup(response.content, 'html.parser')

    def get_product_name(self, soup: BeautifulSoup) -> str:
        """Pobiera nazwę produktu z HTML strony"""
        selectors = ['h1', '.product-title h1', '[itemprop="name"]', '.product-name', '.product-title']
        for selector in selectors:
            element = soup.select_one(selector)
            if element:
                text = element.get_text().strip()
                if text:
                    return text
        return 'Nieznana nazwa'

    def get_product_ean(self, soup: BeautifulSoup) -> str:
        """Pobiera kod EAN ze strony"""
        selectors = [
            '#KodEan strong[itemprop="gtin13"]',
            '#KodEan strong',
            '[itemprop="gtin13"]',
            'div.TbPoz strong[itemprop="gtin13"]'
        ]
        for selector in selectors:
            element = soup.select_one(selector)
            if element:
                value = element.get_text().strip()
                if value:
                    return value
        # Spróbuj znaleźć tekst "Kod EAN:" i odczytać następny element
        text = soup.find(text=re.compile(r'Kod EAN', re.I))
        if text:
            parent = text.parent
            if parent:
                strong = parent.find('strong')
                if strong:
                    value = strong.get_text().strip()
                    if value:
                        return value
        return ''

    def get_previous_price(self, soup: BeautifulSoup) -> str:
        """Pobiera cenę katalogową / przedpromocyjną"""
        selectors = [
            '#CenaPoprzednia strong[content]',
            '#CenaPoprzednia strong',
            'p#CenaPoprzednia strong',
            'div#CenaPoprzednia strong'
        ]
        for selector in selectors:
            element = soup.select_one(selector)
            if element:
                value = element.get_text().strip()
                if value:
                    return value
        text = soup.find(text=re.compile(r'Cena katalogowa|Cena przed', re.I))
        if text and text.parent:
            strong = text.parent.find('strong')
            if strong:
                value = strong.get_text().strip()
                if value:
                    return value
        return ''

    def get_product_images_requests(self, soup: BeautifulSoup, base_url: str) -> list:
        """Pobiera linki do obrazów produktu z załadowanego BeautifulSoup"""
        try:
            images = []
            for img in soup.find_all('img'):
                src = img.get('src') or img.get('data-src') or img.get('data-original') or img.get('data-lazy-src')
                if not src:
                    continue
                src = src.strip()
                if not src or src.startswith('data:') or src.startswith('blob:') or src.startswith('about:blank'):
                    continue
                lower = src.lower()
                if any(x in lower for x in ['logo', 'icon', 'banner', 'sprite', 'placeholder', 'pixel', 'tracking', 'upload-lazy']):
                    continue
                if not re.search(r'\.(jpg|jpeg|png|webp|bmp)(?:[?#]|$)', lower):
                    continue
                if lower.count('/') < 2:
                    continue
                full_url = urljoin(base_url, src)
                if full_url not in images:
                    images.append(full_url)

            logger.info(f"✓ Znaleziono {len(images)} obrazów przez requests")
            return images
        except Exception as e:
            logger.error(f"❌ Błąd pobierania obrazów: {e}")
            return []

    def download_images(self, image_urls: list, product_id: str, download_dir: str = "images"):
        if not image_urls:
            return []

        download_path = Path(download_dir)
        download_path.mkdir(exist_ok=True)

        downloaded_files = []
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/*,*/*;q=0.8',
        }

        for idx, url in enumerate(image_urls, start=1):
            try:
                filename = f"{product_id}-{idx}.jpg"
                filepath = download_path / filename
                response = requests.get(url, headers=headers, timeout=30, stream=True, allow_redirects=True)
                response.raise_for_status()
                content_length = response.headers.get('content-length')
                if content_length and int(content_length) == 0:
                    logger.warning(f"⚠️  Pusty plik: {filename}")
                    continue
                with open(filepath, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                if filepath.stat().st_size == 0:
                    logger.warning(f"⚠️  Plik pusty po zapisie: {filename}")
                    filepath.unlink(missing_ok=True)
                    continue
                downloaded_files.append(str(filepath))
                logger.info(f"✓ Pobrano: {filename}")
            except Exception as e:
                logger.warning(f"❌ Błąd pobierania {url}: {e}")

        return downloaded_files

    def get_product_info(self, url: str) -> dict:
        """Pobiera informacje o produkcie: nazwę, EAN, cenę i obrazy"""
        try:
            logger.info(f"📍 Pobieranie: {url}")
            product_id = self.extract_product_id(url)
            if not product_id:
                logger.warning(f"⚠️  Nie znaleziono ID produktu w URL-u")
                return None

            soup = self.fetch_page_soup(url)
            product_name = self.get_product_name(soup)
            product_ean = self.get_product_ean(soup)
            previous_price = self.get_previous_price(soup)
            image_urls = self.get_product_images_requests(soup, url)

            downloaded_files = []
            image_names = []
            if image_urls:
                downloaded_files = self.download_images(image_urls, product_id)
                image_names = [f"{product_id}-{i+1}" for i in range(len(image_urls))]
                logger.info(f"✓ Pobrano {len(downloaded_files)}/{len(image_urls)} obrazów")
            else:
                logger.warning(f"⚠️  Brak obrazów dla produktu {product_id}")

            return {
                'product_name': product_name,
                'product_id': product_id,
                'product_ean': product_ean,
                'previous_price': previous_price,
                'image_urls': image_urls,
                'downloaded_files': downloaded_files,
                'image_names': image_names,
                'image_count': len(image_urls)
            }
        except Exception as e:
            logger.error(f"❌ Błąd przetwarzania URL-u: {e}")
            return None


def read_urls_from_excel(filepath: str) -> list:
    """Czyta listę URL-i z pliku Excel"""
    try:
        workbook = load_workbook(filepath)
        worksheet = workbook.active
        
        urls = []
        for row in worksheet.iter_rows(values_only=True):
            url = row[0] if row[0] else None
            if url and isinstance(url, str):
                url = url.strip()
                if url.startswith('http'):
                    urls.append(url)
        
        logger.info(f"✓ Wczytano {len(urls)} URL-i z pliku Excel")
        return urls
    
    except FileNotFoundError:
        logger.error(f"❌ Plik nie znaleziony: {filepath}")
        return []
    except Exception as e:
        logger.error(f"❌ Błąd czytania pliku Excel: {e}")
        return []


def save_results_to_excel(results: list, output_filepath: str):
    """Zapisuje wyniki do pliku Excel"""
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Produkty i Obrazy"
        
        # Nagłówki
        worksheet['A1'] = "Nazwa produktu"
        worksheet['B1'] = "ID produktu"
        worksheet['C1'] = "EAN"
        worksheet['D1'] = "Cena przed promocją"
        worksheet['E1'] = "Liczba obrazów"
        worksheet['F1'] = "Nazwy obrazów"
        worksheet['G1'] = "Ścieżki plików"
        worksheet['H1'] = "Linki do obrazów"
        
        # Style nagłówka
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']:
            worksheet[cell].font = header_font
            worksheet[cell].fill = header_fill
        
        # Dane
        for idx, result in enumerate(results, start=2):
            if result:
                worksheet[f'A{idx}'] = result['product_name']
                worksheet[f'B{idx}'] = result['product_id']
                worksheet[f'C{idx}'] = result.get('product_ean', '')
                worksheet[f'D{idx}'] = result.get('previous_price', '')
                worksheet[f'E{idx}'] = result['image_count']
                worksheet[f'F{idx}'] = ', '.join(result['image_names']) if result['image_names'] else 'Brak'
                worksheet[f'G{idx}'] = '\n'.join(result['downloaded_files']) if result['downloaded_files'] else 'Brak'
                worksheet[f'H{idx}'] = '\n'.join(result['image_urls']) if result['image_urls'] else 'Brak'
        
        # Dostosuj szerokość kolumn
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 30
        worksheet.column_dimensions['G'].width = 50
        worksheet.column_dimensions['H'].width = 50
        
        workbook.save(output_filepath)
        logger.info(f"✓ Wyniki zapisane do: {output_filepath}")
        
    except Exception as e:
        logger.error(f"❌ Błąd zapisania wyniku: {e}")


def main():
    print("START: Uruchamiam aplikację...")
    
    print("\n" + "="*60)
    print("  🖼️  POBIERACZ OBRAZÓW PRODUKTÓW - MLAMP.PL")
    print("="*60 + "\n")
    
    # Ścieżki
    script_dir = Path(__file__).parent
    output_file = script_dir / "output.xlsx"

    # Wczytaj plik wejściowy z domyślnych nazw lub z argumentu
    input_file = None
    if len(sys.argv) > 1:
        input_file = Path(sys.argv[1])
        if not input_file.is_absolute():
            input_file = script_dir / input_file
        if not input_file.exists():
            logger.error(f"❌ Plik wejściowy nie istnieje: {input_file}")
            sys.exit(1)
    else:
        candidates = [
            script_dir / "urls.xlsx",
            script_dir / "Linki-do-produktów-MLAMP.xlsx",
            script_dir / "Linki-do-produktów-MLAMP.xlsx",
            script_dir / "input.xlsx",
            script_dir / "links.xlsx"
        ]
        for candidate in candidates:
            if candidate.exists():
                input_file = candidate
                break
        if input_file is None:
            for candidate in script_dir.glob("*.xlsx"):
                if candidate.name.lower() not in {"output.xlsx", "pobierz_obrazy.xlsx"}:
                    input_file = candidate
                    break

    if input_file is None:
        logger.error("❌ Nie znaleziono pliku wejściowego Excel.")
        logger.info("   Umieść plik .xlsx w tym samym folderze co skrypt lub podaj jego nazwę jako argument.")
        sys.exit(1)

    logger.info(f"✓ Używam pliku wejściowego: {input_file.name}")

    # Czytaj URL-i
    urls = read_urls_from_excel(str(input_file))
    if not urls:
        logger.error("❌ Brak URL-i w pliku!")
        sys.exit(1)

    logger.info(f"📌 Do przetworzenia: {len(urls)} produktów\n")

    scraper = ProductImageScraper()
    results = []

    try:
        for idx, url in enumerate(urls, start=1):
            logger.info(f"\n[{idx}/{len(urls)}]")
            result = scraper.get_product_info(url)
            if result:
                results.append(result)
            time.sleep(1)

    except KeyboardInterrupt:
        logger.info("\n⏹️  Przerwano przez użytkownika")
    except Exception as e:
        logger.error(f"\n❌ Błąd: {e}")

    # Zapisz wyniki
    logger.info(f"\n✓ Przetworzono: {len(results)}/{len(urls)} produktów")
    save_results_to_excel(results, str(output_file))

    try:
        logger.info("\n📂 Otwieram plik Excel...")
        subprocess.Popen(['cmd', '/c', f'start "" "{output_file}"'])
    except Exception as e:
        logger.warning(f"⚠️  Nie udało się otworzyć Excela: {e}")
        logger.info(f"📄 Plik znajduje się w: {output_file}")

    logger.info("\n" + "="*60)
    logger.info("✓ GOTOWE!")
    logger.info("="*60)


if __name__ == "__main__":
    main()
