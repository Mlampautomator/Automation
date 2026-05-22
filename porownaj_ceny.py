#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
porownaj_ceny.py  v2
Porownywarka cen: mlamp.pl  vs  ceneo.pl

Jak uzywac:
  python porownaj_ceny.py

Wklej URL produktu z mlamp.pl (np. skopiowany z przegladarki):
  https://mlamp.pl/pl/products/lampa-xeno-52408-saxby-srebna-85659

Skrypt:
  1. Pobiera nazwe, cene regularna i cene po promocji z mlamp.pl
  2. Szuka produktu na ceneo.pl uzywajac numeru katalogowego
  3. Wybiera do 5 ofert TANSZYCH od ceny regularnej mlamp
     (od najdrozszej do najtanszej)
     -> jesli zadnych tanszych nie ma: 3 najblizsze cenowo
  4. Liczy % roznicy wzgledem ceny regularnej mlamp
  5. Zapisuje plik Excel i otwiera go automatycznie
"""

import re
import sys
import json
import time
import subprocess
from datetime import datetime
from urllib.parse import quote_plus, urljoin

# ── Auto-install missing packages ─────────────────────────────────────────────
def _ensure():
    import importlib
    for pkg, mod in [("requests","requests"),("beautifulsoup4","bs4"),
                     ("lxml","lxml"),("openpyxl","openpyxl"),
                     ("selenium","selenium"),("webdriver-manager","webdriver_manager")]:
        try:
            importlib.import_module(mod)
        except ImportError:
            print(f"Instalowanie {pkg}...")
            import subprocess as sp
            sp.check_call([sys.executable,"-m","pip","install",pkg,"-q"])
_ensure()

import requests

# Force UTF-8 output so Polish characters and symbols print correctly on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── HTTP ─────────────────────────────────────────────────────────────────────

BASE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pl-PL,pl;q=0.9,en-US;q=0.8",
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;"
        "q=0.9,image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}


def make_session(referer=None):
    s = requests.Session()
    s.headers.update(BASE_HEADERS)
    if referer:
        s.headers["Referer"] = referer
    a = requests.adapters.HTTPAdapter(pool_connections=4, pool_maxsize=4)
    s.mount("https://", a)
    s.mount("http://",  a)
    return s


def fetch(session, url, retries=2):
    """Return BeautifulSoup or None."""
    for attempt in range(retries + 1):
        try:
            r = session.get(url, timeout=20, allow_redirects=True)
            if r.status_code == 429:
                w = 3 * (attempt + 1)
                print(f"  Rate-limited - czekam {w}s...")
                time.sleep(w)
                continue
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except requests.HTTPError as e:
            code = e.response.status_code if e.response is not None else "?"
            print(f"  HTTP {code}: {url}")
            if code in (403, 404):
                return None
            time.sleep(2 ** attempt)
        except Exception as e:
            if attempt < retries:
                time.sleep(1 + attempt)
            else:
                print(f"  Blad: {e}")
    return None


# ── Price helpers ─────────────────────────────────────────────────────────────

def parse_price(text):
    """'1 299,99 zl' -> 1299.99  |  returns float or None"""
    if not text:
        return None
    s = str(text)
    s = s.replace("\xa0","").replace(" ","").replace(" ","")
    s = s.replace("zl","").replace("zł","").replace("PLN","")
    # Remove thousands separator before converting comma to dot
    # Pattern: digit, dot/comma, exactly 3 digits, then another separator
    s = re.sub(r"(\d)[.,](\d{3})([.,])", lambda m: m.group(1)+m.group(2)+m.group(3), s)
    s = s.replace(",",".")
    m = re.search(r"(\d{1,7}\.?\d{0,2})", s)
    if m:
        try:
            v = float(m.group(1))
            return v if v >= 0.5 else None
        except ValueError:
            pass
    return None


# ═════════════════════════════════════════════════════════════════════════════
# mlamp.pl
# ═════════════════════════════════════════════════════════════════════════════

def _extract_catalog_num(url_or_slug):
    """
    mlamp.pl URLs: /pl/products/[name]-[CATNUM]-[more]-[DBID]
    Returns the first 5-6 digit number embedded in the slug.
    """
    m = re.search(r"-(\d{5,6})-", url_or_slug)
    return m.group(1) if m else None


def _extract_prices_mlamp(soup):
    """Returns (cena_regularna, cena_promo). cena_promo=None if no discount."""

    # 1. JSON-LD with priceSpecification (mlamp.pl uses ListPrice / SalePrice types)
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = next((d for d in data if d.get("@type") == "Product"), None)
            if not data or data.get("@type") != "Product":
                continue
            offers_obj = data.get("offers", {})
            if isinstance(offers_obj, list):
                offers_obj = offers_obj[0] if offers_obj else {}

            # priceSpecification array: ListPrice = regular, SalePrice = promo
            spec = offers_obj.get("priceSpecification", [])
            if isinstance(spec, dict):
                spec = [spec]
            list_price = sale_price = None
            for item in spec:
                ptype = item.get("priceType", "")
                p = parse_price(str(item.get("price", "")))
                if p:
                    if "ListPrice" in ptype:
                        list_price = p
                    elif "SalePrice" in ptype:
                        sale_price = p
            if list_price:
                return list_price, sale_price

            # Single-price offers object
            p = parse_price(str(offers_obj.get("price", "")))
            if p:
                return p, None
        except Exception:
            pass

    # 2. mlamp.pl specific CSS selectors (confirmed via live page inspection)
    #    del.projector_prices__maxprice  = cena regularna (przekreslona)
    #    strong.projector_prices__price  = cena po promocji
    reg_el   = soup.select_one("del.projector_prices__maxprice")
    promo_el = soup.select_one("strong.projector_prices__price")
    if reg_el:
        reg_p   = parse_price(reg_el.get_text())
        promo_p = parse_price(promo_el.get_text()) if promo_el else None
        if reg_p:
            return reg_p, promo_p
    if promo_el and not reg_el:
        p = parse_price(promo_el.get_text())
        if p:
            return p, None

    # 3. Meta tag
    meta_t = (soup.find("meta", {"property": "product:price:amount"}) or
              soup.find("meta", {"name": "product:price:amount"}))
    meta_p = parse_price(meta_t.get("content", "")) if meta_t else None

    # 4. Generic CSS fallbacks
    old_p = None
    cur_p = None

    for sel in ["del", "s", ".price-old", ".old-price", ".price-before",
                ".price-original", "[class*='old-price']",
                "[class*='price-before']", "[class*='original-price']"]:
        for el in soup.select(sel):
            p = parse_price(el.get_text())
            if p and p > 5:
                old_p = p
                break
        if old_p:
            break

    for sel in [".price-current", ".price-sale", ".price-promo",
                "[class*='current-price']", "[class*='sale-price']",
                ".product__price .price", ".product-price", ".price"]:
        for el in soup.select(sel):
            p = parse_price(el.get_text())
            if p and p > 5:
                cur_p = p
                break
        if cur_p:
            break

    if old_p and cur_p and cur_p < old_p:
        return old_p, cur_p
    if old_p:
        return old_p, None
    if cur_p:
        return cur_p, None
    if meta_p:
        return meta_p, None
    return None, None


def _extract_ean_mlamp(soup):
    """
    Extract EAN/GTIN barcode and manufacturer SKU from an mlamp.pl product page.
    Returns (ean, manufacturer_sku) — either or both may be None.
    """
    ean = None
    mfr_sku = None

    # 1. JSON-LD fields: gtin13, gtin8, gtin, sku
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = next((d for d in data if d.get("@type") == "Product"), None)
            if not data or data.get("@type") != "Product":
                continue
            for field in ("gtin13", "gtin8", "gtin14", "gtin"):
                val = str(data.get(field, "")).strip()
                if re.fullmatch(r"\d{8,14}", val):
                    ean = val
                    break
            sku_val = str(data.get("sku", "")).strip()
            if sku_val and sku_val != "None":
                mfr_sku = sku_val
        except Exception:
            pass
        if ean:
            break

    # 2. Product specification table / dl rows
    for row in soup.select("tr, dl dt, .spec-row, .product-params__item"):
        label = row.get_text(" ", strip=True).lower()
        # Look for EAN / GTIN / kod kreskowy
        if any(k in label for k in ("ean", "gtin", "kod kresk", "barcode")):
            nxt = row.find_next_sibling()
            val = (nxt.get_text(strip=True) if nxt else "").strip()
            if not val:
                # value might be in the same cell after a colon
                parts = label.split(":")
                val = parts[-1].strip() if len(parts) > 1 else ""
            if re.fullmatch(r"\d{8,14}", val):
                ean = ean or val
        # Look for nr katalogowy producenta / sku producenta
        if any(k in label for k in ("nr katalogowy producenta", "sku", "kod producenta", "model")):
            nxt = row.find_next_sibling()
            val = (nxt.get_text(strip=True) if nxt else "").strip()
            if val:
                mfr_sku = mfr_sku or val

    # 3. Meta tags
    for attr in ({"property": "product:retailer_item_id"},
                 {"name": "product:retailer_item_id"}):
        tag = soup.find("meta", attr)
        if tag:
            v = tag.get("content", "").strip()
            if re.fullmatch(r"\d{8,14}", v):
                ean = ean or v

    return ean, mfr_sku


def scrape_mlamp(url):
    """
    Scrape product data from a mlamp.pl product URL.
    Returns dict or None.
    """
    session = make_session()
    print(f"\n[1/3] mlamp.pl")
    print(f"  URL: {url}")

    soup = fetch(session, url)
    if not soup:
        print("  BLAD: nie mozna pobrac strony")
        return None

    h1 = soup.find("h1")
    name = h1.get_text(strip=True) if h1 else "Produkt"

    cena, cena_promo = _extract_prices_mlamp(soup)
    catalog_num = _extract_catalog_num(url)
    ean, mfr_sku = _extract_ean_mlamp(soup)

    print(f"  Nazwa:         {name[:70]}")
    print(f"  Nr katalogowy: {catalog_num}")
    if ean:
        print(f"  EAN:           {ean}")
    if mfr_sku:
        print(f"  SKU prod.:     {mfr_sku}")
    print(f"  Cena:          {cena} zl")
    print(f"  Cena po promo: {cena_promo} zl")

    return {
        "name":        name,
        "cena":        cena,
        "cena_promo":  cena_promo,
        "url":         url,
        "catalog_num": catalog_num,
        "ean":         ean,
        "mfr_sku":     mfr_sku,
    }


# ═════════════════════════════════════════════════════════════════════════════
# ceneo.pl
# ═════════════════════════════════════════════════════════════════════════════

# Generic Polish words that should not be used as distinctive search terms
_PL_STOPWORDS = {
    "lampa","lampy","oprawa","oprawy","zewnetrzna","wewnetrzna","scienna",
    "sufitowa","podtynkowa","natynkowa","wiszaca","stojaca","led",
    "czujnik","ruchu","tuba","biala","czarna","szara","zlota","srebrna",
    "aluminium","metalowa","szklana","z","do","ze","i","na","w","przy",
    "downlight","spotlight","reflektor","panel","plafoniera","plafon",
    "zwis","kinkiet","oczko","obrotowa","kierunkowa","regulacja","regulacja",
    "regulacja","regulacją","z","ze","do","na","w","przy","nad","pod",
    "barwa","ciepla","zimna","neutralna","bialy","czarny","szary","inox",
    "stal","nierdzewna","tworzywo","sztuczne","ip","gx","gu","e27","e14",
    "max","min","szt","komplet","zestaw","seria","typ","model","kolor",
}


def _build_search_queries(product_name, catalog_num, ean=None, mfr_sku=None):
    """
    Build ceneo search queries from most to least specific.

    Order of preference:
    1. EAN barcode — uniquely identifies the exact product across all shops
    2. Manufacturer SKU — if different from the mlamp internal catalog number
    3. Model name + brand (words around the catalog number in the product name)
    4. Model name + catalog number
    5. All distinctive words from the product name
    """
    words = product_name.split()
    queries = []

    # 1. EAN — most specific, finds the exact item
    if ean:
        queries.append(ean)

    # 2. Manufacturer SKU (only if it looks like a real code, not just a number we already have)
    if mfr_sku and mfr_sku != catalog_num:
        queries.append(mfr_sku)

    def _distinctive(word_list):
        return [w for w in word_list
                if w.lower() not in _PL_STOPWORDS and len(w) >= 2
                and not re.fullmatch(r"\d+", w)]

    # Locate the catalog number inside the word list
    cat_idx = None
    if catalog_num:
        for i, w in enumerate(words):
            if catalog_num in re.sub(r"[^\d]", "", w):
                cat_idx = i
                break

    if cat_idx is not None:
        before = _distinctive(words[max(0, cat_idx - 3):cat_idx])
        after  = _distinctive(words[cat_idx + 1: cat_idx + 4])

        # 3. Model name (before) + brand name (after) — no internal number
        context = before[-2:] + after[:2]
        if len(context) >= 2:
            queries.append(" ".join(context))
        elif context:
            queries.append(context[0])

        # 4. Model name + catalog number
        if before:
            q = " ".join(before[-2:] + [catalog_num])
            if q not in queries:
                queries.append(q)

    # 5. All distinctive words from the full product name
    meaningful = _distinctive(
        [w for w in words if not (catalog_num and catalog_num in re.sub(r"[^\d]", "", w))]
    )[:5]
    if meaningful:
        q = " ".join(meaningful)
        if q not in queries:
            queries.append(q)

    # Deduplicate while preserving order
    seen, result = set(), []
    for q in queries:
        q = q.strip()
        if q and q not in seen:
            seen.add(q)
            result.append(q)
    return result


def _name_similarity(name_a, name_b):
    """
    Fraction of distinctive tokens from name_a that appear in name_b.
    Numbers (model codes) are included and carry double weight — a mismatch
    like 52408 vs 52409 should decisively lower the score.
    """
    def key_tokens(s):
        tokens = re.sub(r"[^\w\s]", " ", s).lower().split()
        words   = {t for t in tokens if t not in _PL_STOPWORDS and len(t) >= 3
                   and not re.fullmatch(r"\d+", t)}
        numbers = {t for t in tokens if re.fullmatch(r"\d{4,}", t)}  # model numbers
        return words, numbers

    wa, na = key_tokens(name_a)
    wb, nb = key_tokens(name_b)

    if not wa and not na:
        return 0.0

    # Word overlap (weight 1 each)
    word_hits = len(wa & wb)
    # Number overlap (weight 2 each — matching the model number matters a lot)
    num_hits  = len(na & nb) * 2
    # Penalty: numbers present in name_a that don't appear in name_b
    num_miss  = len(na - nb) * 2

    total_possible = len(wa) + len(na) * 2
    score = max(0.0, (word_hits + num_hits - num_miss) / total_possible) if total_possible else 0.0
    return score


def _product_links_with_titles(soup):
    """
    Yield (href, title_text) for every product candidate on a ceneo search page.
    """
    seen = set()
    for a in soup.select("a[href]"):
        href = a.get("href", "")
        is_numeric = re.match(r"^/\d{6,}$", href)
        is_htm     = (re.match(r"^/[A-Za-z][^;]+\.htm$", href)
                      and "/;szukaj" not in href)
        if not (is_numeric or is_htm) or href in seen:
            continue
        seen.add(href)

        # Walk up the DOM to find the nearest product title element
        title = a.get_text(strip=True)
        node  = a.parent
        for _ in range(5):
            if node is None:
                break
            el = node.select_one("h3, h2, .product-name, [class*='product-name'], [class*='name']")
            if el and len(el.get_text(strip=True)) > len(title):
                title = el.get_text(strip=True)
                break
            node = node.parent

        yield href, title


def _best_product_link(soup, mlamp_name, min_score=0.3):
    """
    Return the ceneo product URL that best matches mlamp_name, or None.
    Falls back to the first product link if no candidate clears min_score.
    """
    candidates = list(_product_links_with_titles(soup))
    if not candidates:
        return None

    scored = [((_name_similarity(mlamp_name, title)), href, title)
              for href, title in candidates]
    scored.sort(key=lambda x: x[0], reverse=True)

    best_score, best_href, best_title = scored[0]
    if best_score >= min_score:
        print(f"  Dopasowanie: {best_score:.0%} — {best_title[:60]}")
        return "https://www.ceneo.pl" + best_href

    # No good match — still return the top candidate so caller can keep trying
    return None


def _model_anchors(product_name, catalog_num, ean=None, mfr_sku=None):
    """
    Identify the tokens that pin down THIS specific product across shops:
      - model_codes: the manufacturer model number/code(s) — e.g. 52408, 1818,
                     2310410103, G2521013, POLGU10. The page SKU is authoritative;
                     codes may also be read from the product name.
      - model_names: the distinctive model-name word(s) sitting just before the
                     code in the product name (e.g. 'xeno', 'arezzo', 'pitcher').

    Brand words ('saxby'/'endon') and generic descriptors are deliberately
    ignored — the same lamp is resold under several brand labels on ceneo. The
    EAN is never treated as a model code (it is matched separately).
    """
    words      = product_name.split()
    ean_digits = re.sub(r"[^\d]", "", ean or "")

    def as_code(tok):
        c = re.sub(r"[^A-Za-z0-9]", "", tok)
        # pure-digit model code (3-12 digits) but never the EAN itself
        if re.fullmatch(r"\d{3,12}", c) and c != ean_digits:
            return c
        # alphanumeric model code: letters+digits, >=5 chars
        if (len(c) >= 5 and re.search(r"\d", c) and re.search(r"[A-Za-z]", c)
                and c.lower() not in _PL_STOPWORDS):
            return c
        return None

    # Locate a code token inside the NAME (its position tells us the model name).
    code, cat_idx = None, None
    if catalog_num:
        for i, w in enumerate(words):
            if catalog_num in re.sub(r"[^\d]", "", w):
                code, cat_idx = catalog_num, i
                break
    if code is None:
        for i, w in enumerate(words):
            c = as_code(w)
            if c:
                code, cat_idx = c, i
                break

    def is_name_word(w):
        wl = w.lower()
        return (wl not in _PL_STOPWORDS and len(wl) >= 3
                and wl not in ("saxby", "endon")
                and not re.fullmatch(r"\d+", wl) and as_code(w) is None)

    # Model name = distinctive word(s) immediately before the code.
    names = set()
    if cat_idx is not None:
        for w in words[max(0, cat_idx - 2):cat_idx]:
            if is_name_word(w):
                names.add(w.lower())
    if not names:  # fallback: first distinctive, non-brand, non-code word
        for w in words:
            if is_name_word(w):
                names.add(w.lower())
                break

    # Every code worth matching against a ceneo title (SKU is authoritative).
    codes = set()
    for c in (code, catalog_num, re.sub(r"[^A-Za-z0-9]", "", mfr_sku or "")):
        if c and len(c) >= 3 and c != ean_digits:
            codes.add(c)

    return names, codes, ean


def _match_confidence(anchors, title):
    """
    How confidently `title` denotes the SAME product as the mlamp anchors:
      1.0  EAN appears, or BOTH the model code and a model-name word appear
      0.6  only the model code appears  (risk: a coincidental number collision)
      0.4  only a model-name word appears  (code absent — likely another variant)
      0.0  neither
    """
    names, codes, ean = anchors
    t         = title.lower()
    t_compact = re.sub(r"[^a-z0-9]", "", t)
    t_nums    = set(re.findall(r"\d+", t))

    if ean and ean in t_compact:
        return 1.0

    code_hit = False
    for c in codes:
        if re.fullmatch(r"\d+", c):
            if c in t_nums:                       # exact number, not a substring
                code_hit = True
                break
        elif c.lower() in t_compact:              # alphanumeric model code
            code_hit = True
            break

    name_hit = any(n in t for n in names) if names else False

    if code_hit and name_hit:
        return 1.0
    if code_hit:
        return 0.6
    if name_hit:
        return 0.4
    return 0.0


def _find_ceneo_product(session, product_name, catalog_num, ean=None, mfr_sku=None):
    """
    Search ceneo.pl and return (product_page_url, search_url).

    Accepts a result only when it is confidently the same product — i.e. the
    manufacturer model code AND the model-name word both appear in the ceneo
    title (or the EAN matches). A bare number collision or a generic-keyword
    lookalike is rejected, so 'not on ceneo' items return None instead of
    silently mapping to the wrong product.
    """
    queries  = _build_search_queries(product_name, catalog_num, ean, mfr_sku)
    anchors  = _model_anchors(product_name, catalog_num, ean, mfr_sku)
    names, codes, _ = anchors
    first_search_url = f"https://www.ceneo.pl/;szukaj-{quote_plus(queries[0] if queries else product_name[:30])}"

    print(f"  Identyfikatory: model={sorted(names) or '?'}  kod={sorted(codes) or '?'}")

    best_url, best_conf, best_title = None, 0.0, ""

    for query in queries:
        search_url = f"https://www.ceneo.pl/;szukaj-{quote_plus(query)}"
        print(f"  Szukam ceneo: {search_url}")

        soup = fetch(session, search_url)
        if not soup:
            continue

        # Redirected directly to a product page?
        if soup.select_one(".product-top, h1.product-name"):
            h1    = soup.select_one("h1")
            title = h1.get_text(strip=True) if h1 else ""
            conf  = _match_confidence(anchors, title)
            print(f"  Przekierowanie na produkt (pewnosc {conf:.0%}): {title[:55]}")
            if conf > best_conf:
                best_conf, best_url, best_title = conf, search_url, title
            if conf >= 1.0:
                break
            continue

        # Score every product link on the results page
        for href, title in _product_links_with_titles(soup):
            conf = _match_confidence(anchors, title)
            if conf > best_conf:
                best_conf, best_url, best_title = conf, "https://www.ceneo.pl" + href, title

        if best_conf >= 1.0:
            break

    # Accept only a confident match (model code AND name, or EAN).
    if best_conf >= 1.0:
        print(f"  Dopasowano ({best_conf:.0%}): {best_title[:55]}")
        print(f"  {best_url}")
        return best_url, first_search_url

    if best_url:
        print(f"  Najlepszy kandydat za malo pewny ({best_conf:.0%}): {best_title[:55]}")
    print("  Nie znaleziono pewnego dopasowania na ceneo.pl — podaj URL recznie.")
    return None, first_search_url


def _extract_shop_name(item):
    """
    Extract the real shop name from a ceneo offer item.
    Priority: img[alt] > .shop-label text (skipping 'Firma' placeholder).
    """
    # img alt is the reliable source
    for img in item.select("img[alt]"):
        alt = img.get("alt", "").strip()
        if alt and alt.lower() not in ("firma", "sklep", "shop", ""):
            return alt

    # .shop-label text as fallback (ceneo uses 'Firma' as placeholder)
    shop_el = item.select_one(".shop-label")
    if shop_el:
        txt = shop_el.get_text(strip=True)
        if txt and txt.lower() not in ("firma", "sklep", "shop"):
            return txt

    return "Sklep"


def _extract_product_id(url):
    """Return numeric ceneo product ID from URL, or None."""
    m = re.search(r"ceneo\.pl/(\d{6,})", url)
    return m.group(1) if m else None


def _parse_offers_from_soup(soup, fallback_url):
    """Extract offer dicts from a ceneo soup object."""
    offers = []

    for item in soup.select("div.product-offer"):
        price = None
        price_el = item.select_one(
            "a.product-price .price-format, "
            "a.product-price span.price, "
            ".product-offer__product__price .price-format"
        )
        if price_el:
            price = parse_price(price_el.get_text())
        if not price:
            for sel in [".offer-default", ".price-format", ".price"]:
                el = item.select_one(sel)
                if el:
                    price = parse_price(el.get_text())
                    if price:
                        break
        if not price:
            continue

        shop = _extract_shop_name(item)

        href = ""
        for link_sel in ["a.go-to-shop", "a.button--primary",
                         "a.product-price", "a.go-to-store"]:
            a = item.select_one(link_sel)
            if a:
                href = a.get("href", "")
                if href and href != "#":
                    break
        if not href or href in ("#", "/"):
            href = fallback_url
        if not href.startswith("http"):
            href = urljoin("https://www.ceneo.pl", href)

        # Skip mlamp.pl's own listings — we already have that price
        if "mlamp" in shop.lower() or "mlamp" in href.lower():
            continue

        offers.append({"sklep": shop[:60], "cena": price, "url": href})

    # Best-offer summary box at the top of the product page
    top = soup.select_one(".product-offer-summary")
    if top:
        p_el = top.select_one(".offer-default, .price-format, .price")
        p    = parse_price(p_el.get_text()) if p_el else None
        shop = _extract_shop_name(top)
        a    = top.select_one("a")
        href = ""
        if a:
            href = a.get("href", "")
            if href and href not in ("#", "/"):
                if not href.startswith("http"):
                    href = urljoin("https://www.ceneo.pl", href)
            else:
                href = fallback_url
        if p and "mlamp" not in shop.lower() and "mlamp" not in href.lower():
            offers.append({"sklep": shop[:60], "cena": p, "url": href})

    # Deduplicate by rounded price
    dedup = {}
    for o in offers:
        key = round(o["cena"], 2)
        if key not in dedup:
            dedup[key] = o
    return list(dedup.values())


def _scrape_offers_selenium(product_url):
    """
    Use a headless Chrome browser to render ceneo's JavaScript and collect all offers.
    Returns list of offer dicts, or None if Selenium is unavailable/fails.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
    except ImportError:
        return None

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(f"user-agent={BASE_HEADERS['User-Agent']}")
    options.add_argument("--lang=pl-PL")
    # Suppress DevTools/USB noise
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    driver = None
    try:
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
        except Exception:
            driver = webdriver.Chrome(options=options)

        print(f"  Selenium: otwieram {product_url}")
        driver.get(product_url)

        # Wait until at least one offer row is present (up to 20 s)
        wait = WebDriverWait(driver, 20)
        try:
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div.product-offer")
            ))
        except Exception:
            pass   # page may not have offers; parse whatever loaded

        # Scroll through the page to trigger lazy-loading of additional offers
        for scroll_y in [400, 800, 1200, 1800]:
            driver.execute_script(f"window.scrollTo(0, {scroll_y});")
            time.sleep(0.5)

        # Scroll the offers container into view if present
        try:
            container = driver.find_element(
                By.CSS_SELECTOR,
                "ul.product-offers__list, .js_async-offers-container-standard"
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", container)
        except Exception:
            pass

        # Wait up to 8 s for offer count to stabilise
        prev_count = 0
        for _ in range(8):
            time.sleep(1)
            cur = driver.find_elements(By.CSS_SELECTOR, "div.product-offer")
            if len(cur) == prev_count and len(cur) > 0:
                break
            prev_count = len(cur)

        soup = BeautifulSoup(driver.page_source, "lxml")
        offers = _parse_offers_from_soup(soup, product_url)
        print(f"  Selenium: znaleziono {len(offers)} ofert")
        return offers

    except Exception as e:
        print(f"  Selenium blad: {e}")
        return None
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def _scrape_offers(session, product_url):
    """Scrape seller offers from a ceneo.pl product page."""
    soup = fetch(session, product_url)
    if soup is None:
        return []

    offers = _parse_offers_from_soup(soup, product_url)

    # ceneo loads most offers via JavaScript (js_async-offers-container-standard).
    # When static HTML yields only 0-1 offers, try alternate URLs then Selenium.
    if len(offers) <= 1:
        pid = _extract_product_id(product_url)
        alt_urls = []

        # Check if the async container advertises a load URL via data attributes
        container = soup.select_one(
            "ul.product-offers__list, "
            ".js_async-offers-container-standard, "
            "[data-offers-url], [data-ajax-url]"
        )
        if container:
            for attr in ["data-offers-url", "data-ajax-url", "data-url",
                         "data-load-url", "data-href"]:
                val = container.get(attr, "").strip()
                if val:
                    alt_urls.append(urljoin("https://www.ceneo.pl", val))
                    break

        if pid:
            alt_urls += [
                f"https://www.ceneo.pl/{pid}/oferty",
                f"https://www.ceneo.pl/{pid}/oferty-sklepow",
            ]

        for url in alt_urls:
            if url == product_url:
                continue
            print(f"  Probuje zaladowac oferty: {url}")
            alt_soup = fetch(session, url)
            if not alt_soup:
                continue
            alt_offers = _parse_offers_from_soup(alt_soup, url)
            if len(alt_offers) > len(offers):
                offers = alt_offers
                print(f"  Znaleziono {len(offers)} ofert (statycznie)")
                break

        # Still ≤1 offer — fall back to Selenium for JS rendering
        if len(offers) <= 1:
            print("  Oferty zaladowane przez JS - uruchamiam przegladarke...")
            selenium_offers = _scrape_offers_selenium(product_url)
            if selenium_offers and len(selenium_offers) > len(offers):
                offers = selenium_offers

    # Final filter: remove any mlamp.pl offers (we already have that price)
    offers = [o for o in offers
              if "mlamp" not in o.get("sklep", "").lower()
              and "mlamp" not in o.get("url", "").lower()]

    return offers


def scrape_ceneo(product_name, catalog_num, ean=None, mfr_sku=None, force_url=None):
    """Main ceneo entry point. Returns (offers, ceneo_search_url, ceneo_title)."""
    session = make_session(referer="https://www.ceneo.pl/")

    if force_url:
        product_url = force_url
        search_url  = force_url
    else:
        print(f"\n[2/3] ceneo.pl - nr katalogowy: {catalog_num or '?'}")
        if ean:
            print(f"  EAN: {ean}")
        product_url, search_url = _find_ceneo_product(session, product_name, catalog_num, ean, mfr_sku)
    if not product_url:
        print("  Nie znaleziono strony produktu na ceneo.pl")
        return [], search_url

    print("\n  *** duplicate item found ***")
    print(f"  Strona produktu: {product_url}")

    ceneo_title = ""
    page = fetch(session, product_url)
    if page:
        h1 = page.select_one("h1")
        if h1:
            ceneo_title = h1.get_text(strip=True)
            conf = _match_confidence(_model_anchors(product_name, catalog_num, ean, mfr_sku), ceneo_title)
            print(f"  Tytul ceneo:  {ceneo_title[:70]}")
            print(f"  Pewnosc:      {conf:.0%}")
        offers = _scrape_offers(session, product_url)
    else:
        offers = []

    if offers:
        print(f"  Znaleziono {len(offers)} ofert (bez mlamp.pl)")
    else:
        print("  Brak ofert innych sklepow na ceneo (tylko mlamp.pl lub brak)")
    return offers, product_url, ceneo_title


# ═════════════════════════════════════════════════════════════════════════════
# Offer selection
# ═════════════════════════════════════════════════════════════════════════════

def select_offers(all_offers, ref_price):
    """
    Normal:   up to 5 offers with cena < ref_price, sorted biggest -> smallest.
    Fallback: 3 closest offers when none are cheaper.
    Returns (selected, is_fallback).
    """
    cheaper = sorted(
        [o for o in all_offers if o["cena"] < ref_price],
        key=lambda x: x["cena"], reverse=True        # biggest first
    )
    if cheaper:
        return cheaper[:5], False

    closest = sorted(all_offers, key=lambda x: abs(x["cena"] - ref_price))
    return closest[:3], True


# ═════════════════════════════════════════════════════════════════════════════
# Excel
# ═════════════════════════════════════════════════════════════════════════════

def save_excel(mlamp, offers, ceneo_url, ceneo_title, fallback, out_path):

    wb = Workbook()
    ws = wb.active
    ws.title = "Porownanie cen"

    # ── Palette ──────────────────────────────────────────────────────────────
    P = dict(
        dk_blue  = "1F3864",
        md_blue  = "2B579A",
        lt_blue  = "D6E4F7",
        orange   = "C55A11",
        lt_org   = "FCE4D6",
        green_bg = "E2EFDA",
        amber_bg = "FFF2CC",
        white    = "FFFFFF",
        link     = "0563C1",
        grn_txt  = "375623",
        amb_txt  = "7F6000",
        grey_bg  = "F2F2F2",
    )

    def mkfill(c):
        return PatternFill("solid", fgColor=c)

    thin  = Side(style="thin", color="BBBBBB")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ref   = mlamp["cena"]
    now   = datetime.now().strftime("%Y-%m-%d %H:%M")

    def wcell(r, c, val, *, font=None, fill=None, align=None, fmt=None, link=None):
        cell = ws.cell(row=r, column=c, value=val)
        if font:  cell.font        = font
        if fill:  cell.fill        = fill
        if align: cell.alignment   = align
        cell.border = bdr
        if fmt:   cell.number_format = fmt
        if link:
            cell.hyperlink = link
            cell.font = Font(color=P["link"], underline="single", size=10)
        return cell

    # ══════════════════════════════════════════════════════════════════════
    # ROW 1 — Title
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
        value=f"Porownanie cen mlamp.pl vs ceneo.pl  |  {now}")
    c.font      = Font(bold=True, color=P["white"], size=12)
    c.fill      = mkfill(P["dk_blue"])
    c.alignment = CTR
    c.border    = bdr
    ws.row_dimensions[1].height = 26

    # ══════════════════════════════════════════════════════════════════════
    # ROW 2 — mlamp section banner
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1, value="MLAMP.PL")
    c.font      = Font(bold=True, color=P["white"], size=11)
    c.fill      = mkfill(P["md_blue"])
    c.alignment = CTR
    c.border    = bdr
    ws.row_dimensions[2].height = 22

    # ══════════════════════════════════════════════════════════════════════
    # ROW 3 — mlamp column headers
    # ══════════════════════════════════════════════════════════════════════
    mlamp_hdrs = ["Nazwa produktu", "Cena (zl)",
                  "Cena po promocji (zl)", "URL produktu mlamp.pl",
                  "Link do ceneo.pl", ""]
    hdr_font = Font(bold=True, color=P["white"], size=10)
    hdr_fill = mkfill(P["dk_blue"])
    for ci, h in enumerate(mlamp_hdrs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = CTR; c.border = bdr
    ws.row_dimensions[3].height = 20

    # ══════════════════════════════════════════════════════════════════════
    # ROW 4 — mlamp data row
    # ══════════════════════════════════════════════════════════════════════
    bg4 = mkfill(P["lt_blue"])

    c = ws.cell(row=4, column=1, value=mlamp["name"])
    c.font = Font(bold=True, size=10); c.fill = bg4
    c.alignment = LFT; c.border = bdr

    for col, val, fmt in [
        (2, mlamp["cena"],       '#,##0.00 "zl"'),
        (3, mlamp["cena_promo"], '#,##0.00 "zl"'),
    ]:
        display = val if val is not None else "brak"
        c = ws.cell(row=4, column=col, value=display)
        c.font = Font(bold=True, size=11)
        c.fill = bg4; c.alignment = CTR; c.border = bdr
        if isinstance(display, float):
            c.number_format = fmt

    c = ws.cell(row=4, column=4, value=mlamp["url"])
    c.hyperlink = mlamp["url"]
    c.font = Font(color=P["link"], underline="single", size=10)
    c.fill = bg4; c.alignment = CTR; c.border = bdr

    c = ws.cell(row=4, column=5, value=ceneo_url)
    c.hyperlink = ceneo_url
    c.font = Font(color=P["link"], underline="single", size=10)
    c.fill = bg4; c.alignment = CTR; c.border = bdr

    ws.cell(row=4, column=6).fill = bg4
    ws.cell(row=4, column=6).border = bdr
    ws.row_dimensions[4].height = 36

    # ══════════════════════════════════════════════════════════════════════
    # ROW 5 — spacer
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[5].height = 8

    # ══════════════════════════════════════════════════════════════════════
    # ROW 6 — ceneo banner
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A6:F6")
    if fallback:
        banner = "CENEO.PL — brak tanszych ofert, pokazano 3 najblizsze"
    else:
        banner = "CENEO.PL — oferty tansze od ceny mlamp  (od najdrozszej do najtanszej)"
    c = ws.cell(row=6, column=1, value=banner)
    c.font      = Font(bold=True, color=P["white"], size=11)
    c.fill      = mkfill(P["orange"])
    c.alignment = CTR; c.border = bdr
    ws.row_dimensions[6].height = 22

    # ══════════════════════════════════════════════════════════════════════
    # ROW 7 — ceneo column headers
    # ══════════════════════════════════════════════════════════════════════
    ceneo_hdrs = ["Produkt (ceneo)","Cena (zl)","Roznica od ceny regularnej","% roznicy",
                  "Link do oferty",""]
    for ci, h in enumerate(ceneo_hdrs, 1):
        c = ws.cell(row=7, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = CTR; c.border = bdr
    ws.row_dimensions[7].height = 20

    # ══════════════════════════════════════════════════════════════════════
    # ROWS 8+ — ceneo offers
    # ══════════════════════════════════════════════════════════════════════
    next_r = 8

    if not offers:
        ws.merge_cells(f"A{next_r}:F{next_r}")
        c = ws.cell(row=next_r, column=1, value="Brak ofert ceneo.pl")
        c.font = Font(italic=True, size=10)
        c.fill = mkfill(P["grey_bg"]); c.alignment = CTR; c.border = bdr
        ws.row_dimensions[next_r].height = 18
    else:
        for offer in offers:
            r   = next_r; next_r += 1
            p   = offer["cena"]
            diff    = (p - ref) if ref else None
            pct_val = ((p / ref) - 1) * 100 if ref else None

            diff_str = f"{diff:+.2f} zl" if diff is not None else "—"
            pct_str  = f"{pct_val:+.1f}%"  if pct_val is not None else "—"

            cheaper   = diff is not None and diff < 0
            row_fill  = mkfill(P["green_bg"] if cheaper else P["amber_bg"])

            # Col 1 — ceneo product name
            c = ws.cell(row=r, column=1, value=ceneo_title or offer["sklep"])
            c.font = Font(size=10); c.fill = row_fill
            c.alignment = LFT; c.border = bdr

            # Col 2 — cena
            c = ws.cell(row=r, column=2, value=p)
            c.font = Font(bold=True, size=11)
            c.fill = row_fill; c.alignment = CTR; c.border = bdr
            c.number_format = '#,##0.00 "zl"'

            # Col 3 — roznica
            c = ws.cell(row=r, column=3, value=diff_str)
            c.font = Font(size=10)
            c.fill = row_fill; c.alignment = CTR; c.border = bdr

            # Col 4 — %
            pct_color = P["grn_txt"] if cheaper else P["amb_txt"]
            c = ws.cell(row=r, column=4, value=pct_str)
            c.font = Font(bold=True, color=pct_color, size=10)
            c.fill = row_fill; c.alignment = CTR; c.border = bdr

            # Col 5 — link
            if offer.get("url"):
                c = ws.cell(row=r, column=5, value=offer["url"])
                c.hyperlink = offer["url"]
                c.font = Font(color=P["link"], underline="single", size=10)
            else:
                c = ws.cell(row=r, column=5, value="—")
                c.font = Font(size=10)
            c.fill = row_fill; c.alignment = CTR; c.border = bdr

            c = ws.cell(row=r, column=6, value="")
            c.fill = row_fill; c.border = bdr
            ws.row_dimensions[r].height = 18

    # ══════════════════════════════════════════════════════════════════════
    # Column widths & freeze
    # ══════════════════════════════════════════════════════════════════════
    for i, w in enumerate([44, 16, 30, 13, 54, 3], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    wb.save(out_path)


# ═════════════════════════════════════════════════════════════════════════════
# Index lookup — finds product URL in the local mlamp links Excel file
# ═════════════════════════════════════════════════════════════════════════════

# Path to the Excel that holds all mlamp.pl product URLs (same directory as script)
_LINKS_EXCEL = "Linki-do-produktów-MLAMP.xlsx"

def _xl_path():
    import os as _os
    script_dir = _os.path.dirname(_os.path.abspath(
        __file__ if "__file__" in dir() else sys.argv[0]
    ))
    return _os.path.join(script_dir, _LINKS_EXCEL)


_NOT_ON_CENEO = "NIE_MA_NA_CENEO"


def _lookup_excel(mlamp_url: str = None, index: str = None):
    """
    Look up a row in the links Excel by exact mlamp URL or by index pattern.
    Returns (mlamp_url, ceneo_url) — either value may be None.
    ceneo_url == _NOT_ON_CENEO means the user confirmed it is not on ceneo.
    """
    import os
    from openpyxl import load_workbook
    xl = _xl_path()
    if not os.path.exists(xl):
        return None, None

    wb      = load_workbook(xl, read_only=True, data_only=True)
    pattern = re.compile(rf"-{re.escape(index)}-", re.IGNORECASE) if index else None

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            for i, cell in enumerate(cells):
                if not isinstance(cell, str) or "mlamp.pl" not in cell:
                    continue
                matched = (mlamp_url and cell.strip() == mlamp_url.strip()) or \
                          (pattern and pattern.search(cell))
                if matched:
                    ceneo = None
                    for c in cells[i + 1:]:
                        if isinstance(c, str) and c.strip():
                            ceneo = c.strip()
                            break
                    # Discard stale search URLs saved by older versions
                    if ceneo and ";szukaj-" in ceneo:
                        ceneo = None
                    return cell.strip(), ceneo
    return None, None


def _save_to_excel(mlamp_url: str, ceneo_url: str = None):
    """
    Upsert a row: col A = mlamp URL, col B = ceneo URL (or _NOT_ON_CENEO).
    """
    from openpyxl import load_workbook
    xl = _xl_path()
    try:
        wb = load_workbook(xl)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == mlamp_url.strip():
                    ws.cell(row=cell.row, column=cell.column + 1, value=ceneo_url or "")
                    wb.save(xl)
                    return
        ws.append([mlamp_url, ceneo_url or ""])
        wb.save(xl)
    except Exception as e:
        print(f"  Nie udalo sie zapisac do {_LINKS_EXCEL}: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 64)
    print("  Porownywarka cen:  mlamp.pl  vs  ceneo.pl")
    print("=" * 64)
    print()
    print("Wpisz indeks produktu (np. 52408) LUB wklej pelny URL mlamp.pl.")
    print()

    raw = input("Indeks lub URL: ").strip()
    if not raw:
        print("Anulowano.")
        sys.exit(0)

    # ── Resolve input to mlamp URL + check for saved ceneo URL ──────────
    if raw.startswith("http") and "mlamp.pl" in raw:
        product_url = raw
        _, saved_ceneo_url = _lookup_excel(mlamp_url=product_url)
    elif re.fullmatch(r"\d{4,7}", raw):
        print(f"  Szukam indeksu {raw} w pliku {_LINKS_EXCEL}...")
        product_url, saved_ceneo_url = _lookup_excel(index=raw)
        if product_url:
            print(f"  Znaleziono: {product_url}")
        else:
            print(f"  Indeks {raw} nie znaleziony w {_LINKS_EXCEL}.")
            print("  Wklej pelny URL produktu z mlamp.pl (lub Enter aby anulowac):")
            fallback_url = input("  URL: ").strip()
            if not fallback_url or not fallback_url.startswith("http"):
                print("Anulowano.")
                sys.exit(0)
            product_url = fallback_url
            _, saved_ceneo_url = _lookup_excel(mlamp_url=product_url)
            if not saved_ceneo_url:
                _save_to_excel(product_url)
    else:
        print("Podaj sam indeks (cyfry, np. 52408) lub pelny URL https://mlamp.pl/...")
        sys.exit(1)

    if saved_ceneo_url:
        print(f"  Zapisany URL ceneo: {saved_ceneo_url}")

    # ── Step 1: mlamp.pl ─────────────────────────────────────────────────
    mlamp = scrape_mlamp(product_url)
    if not mlamp:
        print("\nNie udalo sie pobrac danych z mlamp.pl.")
        sys.exit(1)

    # ── Step 2: ceneo.pl ─────────────────────────────────────────────────
    if saved_ceneo_url == _NOT_ON_CENEO:
        print("\n[2/3] ceneo.pl — produkt oznaczony jako niedostepny na ceneo.")
        ceneo_offers, ceneo_url, ceneo_title = [], "", ""

    elif saved_ceneo_url:
        print(f"\n[2/3] ceneo.pl — uzywa zapisanego URL")
        ceneo_offers, ceneo_url, ceneo_title = scrape_ceneo(
            mlamp["name"], mlamp["catalog_num"],
            ean=mlamp.get("ean"), mfr_sku=mlamp.get("mfr_sku"),
            force_url=saved_ceneo_url
        )
        ceneo_url = saved_ceneo_url

    else:
        ceneo_offers, ceneo_url, ceneo_title = scrape_ceneo(
            mlamp["name"], mlamp["catalog_num"],
            ean=mlamp.get("ean"), mfr_sku=mlamp.get("mfr_sku")
        )

        # ── Confirm with user ─────────────────────────────────────────────
        print()
        if ceneo_url:
            print(f"  Znaleziony produkt ceneo: {ceneo_title or '(brak tytulu)'}")
            print(f"  URL: {ceneo_url}")
        else:
            print("  Nie znaleziono produktu na ceneo.")
        print()
        print("  Enter          = to jest ten sam produkt, zapisz i kontynuuj")
        print("  URL ceneo      = wklej poprawny URL produktu ceneo")
        print("  'nie'          = produktu nie ma na ceneo, zapisz i pomin")
        confirm = input("  > ").strip()

        if confirm.lower() in ("nie", "n", "no", "brak"):
            print("  Zapisano: produkt niedostepny na ceneo.")
            _save_to_excel(product_url, _NOT_ON_CENEO)
            ceneo_offers, ceneo_url, ceneo_title = [], "", ""

        elif confirm.startswith("http") and "ceneo.pl" in confirm:
            ceneo_url = confirm
            print(f"  Uzywam: {ceneo_url}")
            ceneo_offers, ceneo_url, ceneo_title = scrape_ceneo(
                mlamp["name"], mlamp["catalog_num"],
                ean=mlamp.get("ean"), mfr_sku=mlamp.get("mfr_sku"),
                force_url=ceneo_url
            )
            _save_to_excel(product_url, ceneo_url)

        else:
            # Enter — accept whatever was found (even if nothing)
            if ceneo_url:
                _save_to_excel(product_url, ceneo_url)

    # ── Step 3: Select offers ────────────────────────────────────────────
    print(f"\n[3/3] Wybieranie ofert...")
    ref = mlamp["cena"]

    if not ceneo_offers:
        print("  Brak ofert ceneo.pl")
        selected, fallback = [], False
    elif not ref:
        print("  Brak ceny referencyjnej mlamp - zapisuje pierwsze 5 ofert")
        selected, fallback = ceneo_offers[:5], False
    else:
        selected, fallback = select_offers(ceneo_offers, ref)
        if fallback:
            print(f"  Brak tanszych od {ref:.2f} zl -> 3 najblizsze")
        else:
            print(f"  Wybrano {len(selected)} ofert tanszych od {ref:.2f} zl")

    # ── Step 4: Excel ─────────────────────────────────────────────────────
    import os
    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    cat      = mlamp["catalog_num"] or "produkt"
    out_path = os.path.abspath(f"porownanie_{cat}_{ts}.xlsx")

    save_excel(mlamp, selected, ceneo_url, ceneo_title, fallback, out_path)
    print(f"\nZapisano: {out_path}")

    try:
        os.startfile(out_path)
    except Exception:
        pass

    print("Gotowe!")


if __name__ == "__main__":
    main()
