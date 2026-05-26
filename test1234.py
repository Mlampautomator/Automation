#!/usr/bin/env python3
"""
Aplikacja do pobierania obrazów produktów z mlamp.pl
Czyta listę URL-i z Excel, pobiera obrazy i zapisuje wyniki.
"""
 
import subprocess
import sys
import os
import time
import re
from urllib.parse import urljoin
from pathlib import Path
 
def install_packages():
    packages = ['openpyxl', 'requests', 'beautifulsoup4']
    for package in packages:
        try:
            __import__(package.replace('-', ''))
        except ImportError:
            print(f"Instalowanie {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])
    print("✓ Wszystkie pakiety są gotowe!\n")
 
try:
    install_packages()
except Exception as e:
    print(f"❌ Błąd instalacji pakietów: {e}")
    sys.exit(1)
 
from openpyxl import Workbook, load_workbook
import logging
import requests
from bs4 import BeautifulSoup
 
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)
 
 
class ProductImageScraper:
    def __init__(self):
        pass
 
    def extract_product_id(self, url: str) -> str:
        match = re.search(r'-(\d{5,6})(?:-|$)', url)
        if match:
            return match.group(1)
        return None
 
    def get_product_name_requests(self, url: str) -> str:
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            for selector in ['h1', '.product-title h1', '[itemprop="name"]', '.productTitle']:
                element = soup.select_one(selector)
                if element:
                    text = element.get_text().strip()
                    if text:
                        return text
            return "Nieznana nazwa"
        except Exception as e:
            logger.debug(f"Błąd pobierania nazwy: {e}")
            return "Nieznana nazwa"
 
    def get_product_images_requests(self, url: str) -> list:
        """
        Na mlamp.pl zdjecia produktu maja w URL segmenty:
          /pol_ps_  -> galeria produktu (product shot)
          /pol_pm_  -> miniaturka produktu (product main)
 
        Wszystko inne to smieci (logo, banery, ikony, inne produkty).
        """
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
 
            soup = BeautifulSoup(response.content, 'html.parser')
            images = []
 
            for img in soup.find_all('img'):
                src = img.get('src') or img.get('data-src') or img.get('data-original')
                if not src:
                    continue
                if not src.startswith('http') and not src.startswith('/'):
                    continue
 
                full_url = urljoin(url, src)
 
                # Akceptuj TYLKO zdjecia produktu
                if '/pol_ps_' not in full_url and '/pol_pm_' not in full_url:
                    continue
 
                # Odrzuc /pl/products/ - to linki do stron innych produktow, nie obrazy
                if '/pl/products/' in full_url:
                    continue
 
                if full_url not in images:
                    images.append(full_url)
 
            logger.info(f"✓ Znaleziono {len(images)} zdjec produktu")
            return images
 
        except Exception as e:
            logger.error(f"❌ Blad pobierania obrazow: {e}")
            return []
 
    def download_images(self, image_urls: list, product_id: str, download_dir: str = "images") -> list:
        if not image_urls:
            return []
 
        download_path = Path(download_dir)
        download_path.mkdir(exist_ok=True)
        downloaded_files = []
 
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
            'Accept-Language': 'pl-PL,pl;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
        }
 
        for idx, url in enumerate(image_urls, start=1):
            try:
                filename = f"{product_id}-{idx}.jpg"
                filepath = download_path / filename
 
                response = requests.get(url, headers=headers, timeout=30, stream=True, allow_redirects=True)
                response.raise_for_status()
 
                content_type = response.headers.get('content-type', '')
                if content_type and 'image' not in content_type:
                    logger.warning(f"⚠️  Nieobrazowy content-type ({content_type}): {filename}")
                    continue
 
                content_length = response.headers.get('content-length')
                if content_length and int(content_length) < 1000:
                    logger.warning(f"⚠️  Plik za maly wg naglowka ({content_length}B): {filename}")
                    continue
 
                with open(filepath, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
 
                file_size = filepath.stat().st_size
                if file_size < 1000:
                    logger.warning(f"⚠️  Plik za maly po zapisaniu ({file_size}B): {filename}")
                    filepath.unlink()
                    continue
 
                # Sprawdz magiczne bajty
                with open(filepath, 'rb') as f:
                    header = f.read(12)
                is_webp = header[:4] == b'RIFF' and header[8:12] == b'WEBP'
                is_image = (
                    header[:2] == b'\xff\xd8' or
                    header[:4] == b'\x89PNG' or
                    header[:4] == b'GIF8' or
                    is_webp
                )
                if not is_image:
                    logger.warning(f"⚠️  Plik nie jest obrazem (zly naglowek): {filename}")
                    filepath.unlink()
                    continue
 
                downloaded_files.append(str(filepath))
                logger.info(f"✓ Pobrano: {filename} ({file_size:,} B)")
 
            except Exception as e:
                logger.warning(f"❌ Blad pobierania {url}: {e}")
 
        return downloaded_files
 
    def get_product_info(self, url: str) -> dict:
        try:
            logger.info(f"📍 Pobieranie: {url}")
            product_id = self.extract_product_id(url)
            if not product_id:
                logger.warning("⚠️  Nie znaleziono ID produktu w URL-u")
                return None
 
            product_name = self.get_product_name_requests(url)
            image_urls = self.get_product_images_requests(url)
 
            downloaded_files = []
            image_names = []
            if image_urls:
                downloaded_files = self.download_images(image_urls, product_id)
                image_names = [f"{product_id}-{i+1}" for i in range(len(downloaded_files))]
                logger.info(f"✓ Pobrano {len(downloaded_files)}/{len(image_urls)} obrazow")
            else:
                logger.warning(f"⚠️  Brak obrazow dla produktu {product_id}")
 
            return {
                'product_name': product_name,
                'product_id': product_id,
                'image_urls': image_urls,
                'downloaded_files': downloaded_files,
                'image_names': image_names,
                'image_count': len(downloaded_files),
            }
        except Exception as e:
            logger.error(f"❌ Blad przetwarzania URL-u: {e}")
            return None
 
 
def read_urls_from_excel(filepath: str) -> list:
    try:
        workbook = load_workbook(filepath)
        worksheet = workbook.active
        urls = []
        for row in worksheet.iter_rows(values_only=True):
            url = row[0] if row[0] else None
            if url and isinstance(url, str):
                url = url.strip()
                if url.startswith('http'):
                    urls.append(url)
        logger.info(f"✓ Wczytano {len(urls)} URL-i z pliku Excel")
        return urls
    except FileNotFoundError:
        logger.error(f"❌ Plik nie znaleziony: {filepath}")
        return []
    except Exception as e:
        logger.error(f"❌ Blad czytania pliku Excel: {e}")
        return []
 
 
def save_results_to_excel(results: list, output_filepath: str):
    try:
        from openpyxl.styles import Font, PatternFill
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Produkty i Obrazy"
 
        col_headers = ["Nazwa produktu", "ID produktu", "Liczba obrazow", "Nazwy obrazow", "Sciezki plikow", "Linki do obrazow"]
        for col, h in enumerate(col_headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
 
        for idx, result in enumerate(results, start=2):
            if result:
                worksheet.cell(row=idx, column=1, value=result['product_name'])
                worksheet.cell(row=idx, column=2, value=result['product_id'])
                worksheet.cell(row=idx, column=3, value=result['image_count'])
                worksheet.cell(row=idx, column=4, value=', '.join(result['image_names']) if result['image_names'] else 'Brak')
                worksheet.cell(row=idx, column=5, value='\n'.join(result['downloaded_files']) if result['downloaded_files'] else 'Brak')
                worksheet.cell(row=idx, column=6, value='\n'.join(result['image_urls']) if result['image_urls'] else 'Brak')
 
        for col, width in zip('ABCDEF', [40, 15, 15, 30, 50, 50]):
            worksheet.column_dimensions[col].width = width
 
        workbook.save(output_filepath)
        logger.info(f"✓ Wyniki zapisane do: {output_filepath}")
    except Exception as e:
        logger.error(f"❌ Blad zapisania wyniku: {e}")
 
 
def main():
    print("\n" + "="*60)
    print("  POBIERACZ OBRAZOW PRODUKTOW - MLAMP.PL  v3")
    print("="*60 + "\n")
 
    script_dir = Path(__file__).parent
    input_file = script_dir / "urls.xlsx"
    output_file = script_dir / "output.xlsx"
 
    if not input_file.exists():
        logger.error(f"❌ Plik {input_file} nie istnieje!")
        sys.exit(1)
 
    urls = read_urls_from_excel(str(input_file))
    if not urls:
        logger.error("❌ Brak URL-i w pliku!")
        sys.exit(1)
 
    logger.info(f"Do przetworzenia: {len(urls)} produktow\n")
 
    scraper = ProductImageScraper()
    results = []
 
    try:
        for idx, url in enumerate(urls, start=1):
            logger.info(f"\n[{idx}/{len(urls)}]")
            result = scraper.get_product_info(url)
            if result:
                results.append(result)
            time.sleep(1)
    except KeyboardInterrupt:
        logger.info("\nPrzerwano przez uzytkownika")
    except Exception as e:
        logger.error(f"\n❌ Blad: {e}")
 
    logger.info(f"\n✓ Przetworzono: {len(results)}/{len(urls)} produktow")
    save_results_to_excel(results, str(output_file))
    logger.info("\n✓ GOTOWE!")
 
    try:
        subprocess.Popen(['cmd', '/c', f'start "" "{output_file}"'])
        time.sleep(1)
    except Exception:
        logger.info(f"Plik znajduje sie w: {output_file}")
 
 
if __name__ == "__main__":
    main()
 
