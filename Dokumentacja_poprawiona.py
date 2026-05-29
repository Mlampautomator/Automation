#!/usr/bin/env python3
"""
Aplikacja do pobierania obrazów produktów z mlamp.pl  v4 (szybka)
Czyta listę URL-i z Excel, pobiera obrazy równolegle i zapisuje wyniki.
"""
 
import subprocess
import sys
import os
import time
import re
import json
from urllib.parse import urljoin
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
 
def install_packages():
    packages = ['openpyxl', 'requests', 'beautifulsoup4']
    for package in packages:
        try:
            __import__(package.replace('-', ''))
        except ImportError:
            print(f"Instalowanie {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])
    print("✓ Wszystkie pakiety są gotowe!\n")
 
try:
    install_packages()
except Exception as e:
    print(f"❌ Błąd instalacji pakietów: {e}")
    sys.exit(1)
 
from openpyxl import Workbook, load_workbook
import logging
import requests
from bs4 import BeautifulSoup
 
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)
 
# ── Jedna sesja HTTP dla wszystkich żądań (connection pooling) ──────────────
SESSION = requests.Session()
SESSION.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'pl-PL,pl;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
})
# Zwiększ pulę połączeń, żeby równoległe wątki się nie blokowały
adapter = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20)
SESSION.mount('https://', adapter)
SESSION.mount('http://', adapter)
 
# Ile produktów przetwarzamy równolegle
MAX_WORKERS_PRODUCTS = 5
# Ile obrazów pobieramy równolegle dla jednego produktu
MAX_WORKERS_IMAGES = 4
 
 
class ProductImageScraper:

    def __init__(self, download_dir: str = "images"):
        self.download_dir = download_dir

    def extract_product_id(self, url: str) -> str:
        match = re.search(r'-(\d{5,6})(?:-|$)', url)
        return match.group(1) if match else None
 
    def _fetch_soup(self, url: str):
        """Pobiera stronę i zwraca BeautifulSoup. Rzuca wyjątek przy błędzie."""
        response = SESSION.get(url, timeout=20)
        response.raise_for_status()
        return BeautifulSoup(response.content, 'html.parser')
 
    def get_product_name(self, soup) -> str:
        for selector in ['h1', '.product-title h1', '[itemprop="name"]', '.productTitle']:
            element = soup.select_one(selector)
            if element:
                text = element.get_text().strip()
                if text:
                    return text
        return "Nieznana nazwa"
 
    def get_product_images(self, soup, base_url: str) -> list:
        """
        Zwraca wyłącznie pełnowymiarowe obrazy (pol_pl_) z linków galerii.
        Tagi <img> zawierają tylko miniatury (pol_ps_, pol_pm_) — pomijamy je.
        """
        seen = set()
        images = []

        for a in soup.find_all('a', href=True):
            href = a['href']
            full_url = urljoin(base_url, href)
            if '/pol_pl_' in full_url and '/pl/products/' not in full_url:
                if full_url not in seen:
                    seen.add(full_url)
                    images.append(full_url)

        logger.info(f"  ✓ Znaleziono {len(images)} zdjęć (pełny rozmiar pol_pl_)")
        return images
 
    def _download_single(self, url: str, filepath: Path) -> bool:
        """Pobiera jeden obraz. Zwraca True przy sukcesie."""
        try:
            response = SESSION.get(
                url, timeout=20, stream=True, allow_redirects=True,
                headers={'Accept': 'image/avif,image/webp,image/apng,image/*,*/*;q=0.8'}
            )
            response.raise_for_status()
 
            content_type = response.headers.get('content-type', '')
            if content_type and 'image' not in content_type:
                return False
            content_length = response.headers.get('content-length')
            if content_length and int(content_length) < 1000:
                return False
 
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=16384):
                    if chunk:
                        f.write(chunk)
 
            file_size = filepath.stat().st_size
            if file_size < 1000:
                filepath.unlink(missing_ok=True)
                return False
 
            with open(filepath, 'rb') as f:
                header = f.read(12)
            is_webp = header[:4] == b'RIFF' and header[8:12] == b'WEBP'
            if not (header[:2] == b'\xff\xd8' or header[:4] == b'\x89PNG'
                    or header[:4] == b'GIF8' or is_webp):
                filepath.unlink(missing_ok=True)
                return False
 
            logger.info(f"    ✓ {filepath.name} ({file_size:,} B)")
            return True
        except Exception as e:
            logger.warning(f"    ❌ {filepath.name}: {e}")
            filepath.unlink(missing_ok=True)
            return False
 
    def download_images(self, image_urls: list, product_id: str, download_dir: str = None) -> list:
        if not image_urls:
            return []
        download_path = Path(download_dir if download_dir is not None else self.download_dir)
        download_path.mkdir(parents=True, exist_ok=True)
 
        tasks = []
        for idx, url in enumerate(image_urls, 1):
            filepath = download_path / f"{product_id}-{idx}.jpg"
            tasks.append((url, filepath))
 
        downloaded = []
        # Pobieramy obrazy jednego produktu równolegle
        with ThreadPoolExecutor(max_workers=MAX_WORKERS_IMAGES) as executor:
            futures = {executor.submit(self._download_single, url, fp): fp for url, fp in tasks}
            for future in as_completed(futures):
                if future.result():
                    downloaded.append(str(futures[future]))
 
        # Sortuj, żeby kolejność była deterministyczna
        downloaded.sort()
        return downloaded
 
    def get_product_info(self, url: str) -> dict:
        try:
            logger.info(f"📍 {url}")
            product_id = self.extract_product_id(url)
            if not product_id:
                logger.warning("  ⚠️  Brak ID produktu w URL")
                return None
 
            soup = self._fetch_soup(url)
            product_name = self.get_product_name(soup)
            image_urls = self.get_product_images(soup, url)
 
            downloaded_files = []
            if image_urls:
                downloaded_files = self.download_images(image_urls, product_id)
                logger.info(f"  ✓ Pobrano {len(downloaded_files)}/{len(image_urls)} obrazów")
            else:
                logger.warning(f"  ⚠️  Brak obrazów dla {product_id}")
 
            image_names = [Path(f).stem for f in downloaded_files]
            return {
                'product_name': product_name,
                'product_id': product_id,
                'product_url': url,
                'image_urls': image_urls,
                'downloaded_files': downloaded_files,
                'image_names': image_names,
                'image_count': len(downloaded_files),
            }
        except Exception as e:
            logger.error(f"  ❌ Błąd: {e}")
            return None
 
 
def read_urls_from_excel(filepath: str) -> list:
    """
    Czyta URL-e z pliku linków mlamp.pl.
    Obsługuje dwa formaty:
      - Nowy (linki_mlamp_*.xlsx): kolumna D (index 3) = URL mlamp.pl
      - Stary (urls.xlsx itp.):    kolumna A (index 0) = URL
    """
    try:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        worksheet = workbook.active
        urls = []
        for row in worksheet.iter_rows(values_only=True):
            if not row:
                continue
            # Nowy format: kolumna D (index 3) zawiera URL mlamp.pl
            if len(row) >= 4:
                cell = row[3]
                if cell and isinstance(cell, str) and cell.strip().startswith('https://mlamp.pl'):
                    urls.append(cell.strip())
                    continue
            # Fallback: kolumna A
            cell = row[0]
            if cell and isinstance(cell, str) and cell.strip().startswith('http'):
                urls.append(cell.strip())
        logger.info(f"✓ Wczytano {len(urls)} URL-i")
        return urls
    except FileNotFoundError:
        logger.error(f"❌ Plik nie znaleziony: {filepath}")
        return []
    except Exception as e:
        logger.error(f"❌ Błąd czytania Excel: {e}")
        return []
 
 
def save_results_to_excel(results: list, output_filepath: str):
    try:
        from openpyxl.styles import Font, PatternFill
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Produkty i Obrazy"
 
        headers = ["Nazwa produktu", "ID produktu", "Liczba obrazów",
                   "Nazwy obrazów", "Link do produktu", "Linki do obrazów"]
        for col, h in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
 
        for idx, result in enumerate(results, 2):
            if result:
                worksheet.cell(row=idx, column=1, value=result['product_name'])
                worksheet.cell(row=idx, column=2, value=result['product_id'])
                worksheet.cell(row=idx, column=3, value=result['image_count'])
                worksheet.cell(row=idx, column=4, value=', '.join(result['image_names']) or 'Brak')
                worksheet.cell(row=idx, column=5, value=result['product_url'])
                worksheet.cell(row=idx, column=6, value='\n'.join(result['image_urls']) or 'Brak')
 
        for col, width in zip('ABCDEF', [40, 15, 15, 30, 50, 50]):
            worksheet.column_dimensions[col].width = width
 
        workbook.save(output_filepath)
        logger.info(f"✓ Wyniki zapisane do: {output_filepath}")
    except Exception as e:
        logger.error(f"❌ Błąd zapisu: {e}")
 
 
def _load_log(log_path: str) -> dict:
    try:
        with open(log_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_log(log_path: str, log: dict):
    with open(log_path, 'w', encoding='utf-8') as f:
        json.dump(log, f, ensure_ascii=False, indent=2)


def main():
    print("\n" + "=" * 60)
    print("  POBIERACZ OBRAZOW PRODUKTOW - MLAMP.PL  v4 (szybki)")
    print("=" * 60 + "\n")

    script_dir = Path(__file__).parent
    input_file = script_dir / "Linki-do-produktow-MLAMP.xlsx"
    # Also accept the Polish filename with diacritics
    if not input_file.exists():
        input_file = script_dir / "Linki-do-produktów-MLAMP.xlsx"
    output_file = script_dir / "output.xlsx"
    log_path    = str(script_dir / "dokumentacja_log.json")
    images_dir  = str(script_dir / "images")

    if not input_file.exists():
        logger.error(f"Plik {input_file} nie istnieje!")
        sys.exit(1)

    logger.info(f"Plik wejsciowy: {input_file.name}")

    urls = read_urls_from_excel(str(input_file))
    if not urls:
        logger.error("Brak URL-i!")
        sys.exit(1)

    # Load download log — skip products already on disk
    log = _load_log(log_path)

    to_fetch  = []   # (original_index, url)
    skip_map  = {}   # original_index -> cached result

    for i, url in enumerate(urls):
        if url in log:
            cached = log[url]
            files  = cached.get('downloaded_files', [])
            # Accept from cache only if at least one image file still exists
            if files and any(Path(f).exists() for f in files):
                skip_map[i] = cached
                continue
        to_fetch.append((i, url))

    logger.info(f"Juz pobrano: {len(skip_map)}  |  Do pobrania: {len(to_fetch)}\n")

    scraper     = ProductImageScraper(download_dir=images_dir)
    new_results = {}   # original_index -> result
    start       = time.time()

    if to_fetch:
        try:
            with ThreadPoolExecutor(max_workers=MAX_WORKERS_PRODUCTS) as executor:
                futures = {
                    executor.submit(scraper.get_product_info, url): (i, url)
                    for i, url in to_fetch
                }
                done = 0
                for future in as_completed(futures):
                    i, url = futures[future]
                    done += 1
                    try:
                        result = future.result()
                    except Exception as e:
                        logger.error(f"Blad {url}: {e}")
                        result = None
                    new_results[i] = result
                    if result:
                        log[url] = result
                    logger.info(f"[{done}/{len(to_fetch)}] gotowe\n")
        except KeyboardInterrupt:
            logger.info("Przerwano przez uzytkownika")

    # Persist log after every run so new downloads are remembered
    _save_log(log_path, log)

    # Rebuild full results in original URL order
    results = [
        skip_map.get(i) or new_results.get(i)
        for i in range(len(urls))
    ]

    elapsed = time.time() - start
    logger.info(f"Przetworzono nowych: {sum(1 for r in new_results.values() if r)}/{len(to_fetch)}")
    logger.info(f"Czas: {elapsed:.1f}s")
    save_results_to_excel(results, str(output_file))
    logger.info("GOTOWE!")
 
    try:
        subprocess.Popen(['cmd', '/c', f'start "" "{output_file}"'])
    except Exception:
        logger.info(f"Plik: {output_file}")
 
 
if __name__ == "__main__":
    main()
